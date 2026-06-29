from __future__ import annotations

import json
import tempfile
import unittest
import asyncio
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from data_master_app.converter import (
    analyze_colors_file,
    analyze_product_model_files,
    analyze_uploaded_file,
    apply_enrichment_session_to_rows,
    apply_typical_products_to_products,
    build_pim_product,
    category_ids,
    convert_products_file,
    convert_colors_file,
    convert_systems_file,
    convert_uploaded_file,
    map_source_row,
    map_system_row,
    read_source_tables,
)
from data_master_app.mapping import (
    PRODUCT_FIELDS,
    SYSTEM_FIELDS,
    apply_column_mapping,
    apply_column_mapping_profile,
    apply_mapping_profile_to_rows,
    field_quality,
    product_fields_from_pim_bundle,
    suggest_mapping,
)
from data_master_app.main import analyze_products_page, app, home, product_model_accept, product_model_files
from data_master_app.building_elements_ui import render_building_elements_home
from data_master_app.colors_ui import render_colors_home
from data_master_app.web_ui import render_home, render_main_menu
from mapping_studio.models import ProductReferenceIndex
from mapping_studio.services.building_preview import convert_building_elements_from_tables, preview_building_elements
from mapping_studio.services.pim_model_loader import load_building_element_model
from mapping_studio.services.source_reader import SourceTable
from starlette.datastructures import UploadFile


class ConverterTests(unittest.TestCase):
    def test_main_menu_links_to_independent_mapping_sections(self):
        html = render_main_menu()

        self.assertIn("Mapowanie Produktów", html)
        self.assertIn('href="/products"', html)
        self.assertIn("Mapowanie Building Elementów", html)
        self.assertIn('href="/building-elements"', html)

    def test_web_ui_has_product_variant_row_rules_and_hidden_rule_rows(self):
        html = render_home()

        self.assertIn("<title>BuildData AI Products</title>", html)
        self.assertIn("<h1>BuildData AI Products</h1>", html)
        self.assertNotIn("systemsForm", html)
        self.assertNotIn("analyzeSystemsBtn", html)
        self.assertNotIn("building_elements.json", html)
        self.assertIn("data-row-rule=\"rowMode\"", html)
        self.assertIn('value="product_variants"', html)
        self.assertIn('data-row-rule="productRowValues" list=', html)
        self.assertIn('data-row-rule="groupRowValues" list=', html)
        self.assertIn('placeholder="Product"', html)
        self.assertIn('placeholder="Article"', html)
        self.assertIn("rowTypeProductValues", html)
        self.assertIn("rowTypeVariantValues", html)
        self.assertNotIn("id=\"addRowRuleBtn\"", html)
        self.assertNotIn("data-remove-row-rule", html)
        self.assertNotIn("data-row-rule=\"groupTargetPath\"", html)
        self.assertNotIn("data-row-rule=\"groupSourceColumn\"", html)
        self.assertNotIn("data-row-rule=\"groupNameColumn\"", html)
        self.assertIn("data-row-rule=\"idColumn\"", html)
        self.assertIn("data-row-rule=\"parentIdColumn\"", html)
        self.assertIn("Kolumna ID produktu", html)
        self.assertIn("Kolumna Parent ID wariantu", html)
        self.assertNotIn("if (rule.id_column) columns.add(rule.id_column)", html)
        self.assertNotIn("if (rule.parent_id_column) columns.add(rule.parent_id_column)", html)
        self.assertNotIn("data-row-rule=\"productNameColumn\"", html)
        self.assertNotIn("data-row-rule=\"variantNameColumn\"", html)
        self.assertIn("product.name.value", html)
        self.assertIn("data-cleanup=\"choiceMap\"", html)
        self.assertIn("data-choice-map-target", html)
        self.assertIn("data-product-type-rule=\"sourceColumn\"", html)
        self.assertIn("data-product-type-rule=\"otherValues\"", html)
        self.assertIn("prodctTypeId", html)
        self.assertNotIn("value=\"custom_attribute\"", html)
        self.assertNotIn("value=\"type_series[].additional_properties[]\"", html)
        self.assertNotIn("data-cleanup=\"customName\"", html)
        self.assertNotIn("data-cleanup=\"featureGroup\"", html)
        self.assertIn("data-cleanup=\"unitConversionFactor\"", html)
        self.assertIn("data-cleanup=\"targetUnit\"", html)
        self.assertIn("id=\"productModelsFile\"", html)
        self.assertIn("id=\"productAttributesFile\"", html)
        self.assertNotIn("id=\"productProductsFile\"", html)
        self.assertIn("action=\"/product-model-accept\"", html)
        self.assertIn("name=\"products_models_file\"", html)
        self.assertIn("name=\"products_attributes_file\"", html)
        self.assertNotIn("name=\"products_file\"", html)
        self.assertIn("id=\"acceptProductModelBtn\" data-i18n=\"model.accept\" disabled", html)
        self.assertIn("window.productModelFileChanged", html)
        self.assertNotIn("id=\"productModelFile\"", html)
        self.assertNotIn("new File([item.file]", html)
        self.assertIn("window.acceptProductModelSubmit", html)
        self.assertIn("/product-model", html)
        self.assertIn("/product-model-files/", html)
        self.assertIn("productModelFilesForProject", html)
        self.assertIn("renderProductModelPreview", html)
        self.assertIn(".mapping-row.rule-owned-row", html)
        self.assertIn("rowRuleTargetPaths", html)
        self.assertIn("legacyProductMapping", html)
        self.assertIn("productMappingsByModel[activeProductRootModelId]", html)
        self.assertIn("split(/[;\\n|,]+/)", html)
        self.assertNotIn("split(/[;\n|,]+/)", html)

    def test_colors_ui_has_choice_mapping_editor(self):
        html = render_colors_home()

        self.assertIn("data-color-choice-field", html)
        self.assertIn("color_choice_mapping", html)

    def test_building_elements_ui_explains_product_identity(self):
        html = render_building_elements_home()

        self.assertIn("identyfikuje produkt w warstwie", html)
        self.assertIn("Referencyjne products.json jest opcjonalne", html)
        self.assertIn('id="elementProductModelsFile"', html)
        self.assertIn('id="elementProductAttributesFile"', html)
        self.assertIn('id="elementProductIdentityFieldSelect"', html)
        self.assertIn("_product_identity", html)
        self.assertIn("kilka produktów w jednej komórce", html)
        self.assertIn("localStorage.setItem(ELEMENT_WORKSPACE_KEY", html)
        self.assertIn("return currentElementMapping", html)
        self.assertIn("generateBuildingElements", html)
        self.assertIn("building_elements.json", html)
        self.assertIn("elementWorkflowMode", html)
        self.assertIn("elementLivePreview", html)
        self.assertIn("updateElementLivePreview", html)
        self.assertIn("setElementPreviewIndex", html)
        self.assertIn("Element ${elementPreviewIndex + 1} / ${systems.length}", html)
        self.assertIn("lastElementPreview?.systems || []", html)
        self.assertIn("Podgląd na żywo mapowania", html)
        self.assertIn("modelBuilderPanel", html)
        self.assertIn("Stwórz własne systemy z modelu", html)
        self.assertIn("modelBuilderRows", html)
        self.assertIn("modelBuilderMappingProfile", html)
        self.assertIn("Dodaj obiekt z modelu", html)
        self.assertNotIn("modelBuilderEditorInline", html)
        self.assertIn("data-model-builder-input", html)
        self.assertIn("data-model-builder-level-id", html)
        self.assertIn("data-model-builder-parent-id", html)
        self.assertIn("__level_id__", html)
        self.assertIn("__parent_id__", html)
        self.assertIn("ID parenta", html)
        self.assertIn("Kolumna ID tego levela", html)
        self.assertNotIn("Level name z mapowania", html)

    def test_building_preview_splits_comma_separated_layer_products(self):
        product_a = {"Id": 101}
        product_b = {"Id": 102}
        product_index = ProductReferenceIndex(
            products_count=2,
            by_code={"sku1": product_a, "sku2": product_b},
        )

        result = preview_building_elements(
            [
                {
                    "System": "S1",
                    "Wariant": "W1",
                    "Warstwa": "L1",
                    "Produkt": "SKU1, SKU2",
                }
            ],
            {},
            product_index,
        )

        products = result["systems"][0]["variants"][0]["layers"][0]["products"]
        self.assertEqual([item["raw"] for item in products], ["SKU1", "SKU2"])
        self.assertEqual([item["product_id"] for item in products], [101, 102])
        self.assertEqual([item["identity_source"] for item in products], ["code", "code"])

    def test_building_elements_convert_writes_json_from_profile(self):
        model = load_building_element_model(
            {
                "buildingElementsModels.json": b"""{
                  "models": [
                    {"Id": 74, "Name": "Building Element", "modelType": "Building_Element"},
                    {"Id": 75, "Name": "Variant", "modelType": "Attribute"},
                    {"Id": 76, "Name": "Layer", "modelType": "Attribute"}
                  ]
                }""",
                "buildingElementsAttributes.json": b"""{
                  "attributes": [
                    {"Id": 280, "ProductModelId": 74, "AttributeName": "name", "DispName": "Name", "AttributeType": "VarChar", "deleted": false},
                    {"Id": 283, "ProductModelId": 74, "AttributeName": "variants", "DispName": "Variants", "AttributeType": "Model_Array", "TargetModelId": 75, "deleted": false},
                    {"Id": 284, "ProductModelId": 75, "AttributeName": "variant_name", "DispName": "Variant Name", "AttributeType": "VarChar", "deleted": false},
                    {"Id": 285, "ProductModelId": 75, "AttributeName": "layers", "DispName": "Layers", "AttributeType": "Model_Array", "TargetModelId": 76, "deleted": false},
                    {"Id": 287, "ProductModelId": 76, "AttributeName": "layer_name", "DispName": "Layer name", "AttributeType": "VarChar", "deleted": false}
                  ]
                }""",
            }
        )
        tables = [SourceTable("Systemy", [{"System": "S1", "Wariant": "W1", "Warstwa": "L1"}])]
        profile = {
            "_levels": {"model.74": {"level_name_field": "building_element.name.value"}},
            "building_element.name.value": {"table": "Systemy", "column": "System", "cleanup": {"trim": True}},
            "building_element.variant_name.value": {"table": "Systemy", "column": "Wariant", "cleanup": {"trim": True}},
            "building_element.layer_name.value": {"table": "Systemy", "column": "Warstwa", "cleanup": {"trim": True}},
        }

        with tempfile.TemporaryDirectory() as tmp:
            result = convert_building_elements_from_tables("systems.json", b"{}", tables, profile, model, None, Path(tmp))
            payload = json.loads((Path(tmp) / result["job_id"] / "building_elements.json").read_text(encoding="utf-8"))

        self.assertEqual(result["building_elements_count"], 1)
        attrs = payload["buildingElements"][0]["dataVersions"][0]["productAttributes"]
        self.assertTrue(any(attr["AttributeId"] == 280 and attr["varcharValue"] == "S1" for attr in attrs))
        self.assertTrue(any(attr["AttributeId"] == 284 and attr["ParentAttributeId"] == 283 for attr in attrs))
        self.assertTrue(any(attr["AttributeId"] == 287 and attr["ParentAttributeId"] == 285 for attr in attrs))

    def test_building_elements_convert_uses_level_parent_ids_for_nested_hashes(self):
        model = load_building_element_model(
            {
                "buildingElementsModels.json": b"""{
                  "models": [
                    {"Id": 74, "Name": "Building Element", "modelType": "Building_Element"},
                    {"Id": 75, "Name": "Variant", "modelType": "Attribute"},
                    {"Id": 76, "Name": "Layer", "modelType": "Attribute"}
                  ]
                }""",
                "buildingElementsAttributes.json": b"""{
                  "attributes": [
                    {"Id": 280, "ProductModelId": 74, "AttributeName": "name", "DispName": "Name", "AttributeType": "VarChar", "deleted": false},
                    {"Id": 283, "ProductModelId": 74, "AttributeName": "variants", "DispName": "Variants", "AttributeType": "Model_Array", "TargetModelId": 75, "deleted": false},
                    {"Id": 284, "ProductModelId": 75, "AttributeName": "variant_name", "DispName": "Variant Name", "AttributeType": "VarChar", "deleted": false},
                    {"Id": 285, "ProductModelId": 75, "AttributeName": "layers", "DispName": "Layers", "AttributeType": "Model_Array", "TargetModelId": 76, "deleted": false},
                    {"Id": 287, "ProductModelId": 76, "AttributeName": "layer_name", "DispName": "Layer name", "AttributeType": "VarChar", "deleted": false}
                  ]
                }""",
            }
        )
        tables = [
            SourceTable(
                "Systemy",
                [{"SystemId": "S1", "VariantId": "V1", "LayerId": "L1", "System": "System 1", "Wariant": "Wariant 1", "Warstwa": "Warstwa 1"}],
            )
        ]
        profile = {
            "_levels": {
                "model.74": {"table": "Systemy", "id_column": "SystemId"},
                "model.74.attribute.283": {"table": "Systemy", "id_column": "VariantId", "parent_id_column": "SystemId"},
                "model.75.attribute.285": {"table": "Systemy", "id_column": "LayerId", "parent_id_column": "VariantId"},
            },
            "building_element.name.value": {"table": "Systemy", "column": "System", "cleanup": {"trim": True}},
            "building_element.variant_name.value": {"table": "Systemy", "column": "Wariant", "cleanup": {"trim": True}},
            "building_element.layer_name.value": {"table": "Systemy", "column": "Warstwa", "cleanup": {"trim": True}},
        }

        with tempfile.TemporaryDirectory() as tmp:
            result = convert_building_elements_from_tables("systems.json", b"{}", tables, profile, model, None, Path(tmp))
            payload = json.loads((Path(tmp) / result["job_id"] / "building_elements.json").read_text(encoding="utf-8"))

        attrs = payload["buildingElements"][0]["dataVersions"][0]["productAttributes"]
        variant_attr = next(attr for attr in attrs if attr["AttributeId"] == 284)
        layer_attr = next(attr for attr in attrs if attr["AttributeId"] == 287)

        self.assertTrue(variant_attr["hash"])
        self.assertEqual(layer_attr["parentHash"], variant_attr["hash"])

    def test_systems_endpoint_is_removed_from_app(self):
        paths = {route.path for route in app.routes}

        self.assertNotIn("/convert-systems", paths)

    def test_product_model_accept_form_redirects_to_working_page(self):
        response = asyncio.run(
            product_model_accept(
                products_models_file=UploadFile(
                    BytesIO(json.dumps({"models": [{"Id": 41, "Name": "Product", "modelType": "Product"}]}).encode("utf-8")),
                    filename="whatever-a.json",
                ),
                products_attributes_file=UploadFile(
                    BytesIO(
                        json.dumps(
                            {
                                "attributes": [
                                    {
                                        "Id": 116,
                                        "ProductModelId": 41,
                                        "AttributeName": "Nazwa",
                                        "DispName": "Nazwa",
                                        "AttributeType": "TextBox",
                                    }
                                ]
                            }
                        ).encode("utf-8")
                    ),
                    filename="whatever-b.json",
                ),
            )
        )

        self.assertEqual(response.status_code, 303)
        location = response.headers["location"]
        self.assertTrue(location.startswith("/products?product_model_id="))

        page = home(product_model_id=location.split("=", 1)[1])
        body = page.body.decode("utf-8")
        self.assertEqual(page.status_code, 200)
        self.assertIn("INITIAL_PRODUCT_MODEL", body)
        self.assertIn("product.name.value", body)
        self.assertIn("Model produktu został zaakceptowany", body)
        self.assertIn("Wczytane pola modelu", body)
        self.assertIn('id="productModelId" type="hidden" name="product_model_id" value="', body)
        self.assertIn('id="productsProductModelId" type="hidden" name="product_model_id" value="', body)
        self.assertIn('id="productsForm" action="/analyze-products-page" method="post" enctype="multipart/form-data"', body)
        self.assertIn('id="productsFile" name="file" type="file" accept=".xlsx,.xlsm,.json,.csv,.tsv">', body)
        self.assertIn('type="submit" class="secondary" id="analyzeProductsBtn"', body)
        self.assertIn('id="clearProductSessionBtn"', body)
        self.assertIn("Pobierz Excel do akceptacji produktów", body)
        self.assertNotIn('id="productsForm" action="/analyze"', body)

        files_response = product_model_files(location.split("=", 1)[1])
        self.assertEqual({item["name"] for item in files_response["files"]}, {"productsModels.json", "productsAttributes.json"})
        self.assertTrue(all(item["dataUrl"].startswith("data:application/json;base64,") for item in files_response["files"]))

    def test_analyze_products_page_returns_mapping_workspace(self):
        accept_response = asyncio.run(
            product_model_accept(
                products_models_file=UploadFile(
                    BytesIO(json.dumps({"models": [{"Id": 41, "Name": "Product", "modelType": "Product"}]}).encode("utf-8")),
                    filename="productsModels.json",
                ),
                products_attributes_file=UploadFile(
                    BytesIO(
                        json.dumps(
                            {
                                "attributes": [
                                    {
                                        "Id": 116,
                                        "ProductModelId": 41,
                                        "AttributeName": "Nazwa",
                                        "DispName": "Nazwa",
                                        "AttributeType": "TextBox",
                                    }
                                ]
                            }
                        ).encode("utf-8")
                    ),
                    filename="productsAttributes.json",
                ),
            )
        )
        model_id = accept_response.headers["location"].split("=", 1)[1]

        response = asyncio.run(
            analyze_products_page(
                file=UploadFile(BytesIO(json.dumps([{"Nazwa": "Produkt testowy"}]).encode("utf-8")), filename="client.json"),
                product_model_id=model_id,
            )
        )

        body = response.body.decode("utf-8")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Mapping: products", body)
        self.assertIn("Cecha modelu PIM", body)
        self.assertIn("Reguła wiersza i hierarchii", body)
        self.assertIn('data-row-rule="rowMode"', body)
        self.assertIn('data-map-mode="products"', body)
        self.assertIn("client.json", body)
        self.assertIn("INITIAL_ANALYSIS", body)
        self.assertIn('id="productsSourceId" type="hidden" name="products_source_id" value="', body)

    def test_product_model_analysis_can_select_root_model(self):
        files = {
            "productsModels.json": json.dumps(
                {
                    "models": [
                        {"Id": 41, "Name": "Model firmowy", "modelType": "Product"},
                        {"Id": 42, "Name": "Model typowy", "modelType": "Product"},
                    ]
                }
            ).encode("utf-8"),
            "productsAttributes.json": json.dumps(
                {
                    "attributes": [
                        {"Id": 116, "ProductModelId": 41, "AttributeName": "Nazwa", "DispName": "Nazwa", "AttributeType": "TextBox"},
                        {"Id": 117, "ProductModelId": 42, "AttributeName": "Kod", "DispName": "Kod", "AttributeType": "TextBox"},
                    ]
                }
            ).encode("utf-8"),
        }

        result = analyze_product_model_files(files, product_root_model_id=42)

        self.assertEqual(result["selected_root_model_id"], 42)
        self.assertEqual([item["id"] for item in result["product_models"]], [41, 42])
        labels = [field["label"] for field in result["target_fields"]]
        self.assertIn("Kod", labels)
        self.assertNotIn("Nazwa", labels)

    def test_server_rendered_mapping_keeps_type_series_table(self):
        html = render_home(
            {"model_id": "model-1", "target_fields": [{"key": "product.name.value", "label": "Nazwa", "group": "Product identity"}]},
            {
                "mode": "products",
                "source_id": "source-1",
                "filename": "client.json",
                "analysis": {
                    "tables": [
                        {
                            "name": "Sheet1",
                            "rows": 1,
                            "columns": ["Grubość"],
                            "sample_rows": [{"Grubość": "12,5 mm"}],
                            "product_mapping": {
                                "score": 1,
                                "mapping": {"Grubość": "type_series[].thickness.value"},
                                "confidence": {"Grubość": 0.91},
                                "field_quality": {"Grubość": {"missing_rows": 0}},
                                "target_fields": [
                                    {"key": "product.name.value", "label": "Nazwa", "group": "Product identity"},
                                    {"key": "type_series[].thickness.value", "label": "Grubość", "group": "Type series"},
                                ],
                            },
                        }
                    ]
                },
            },
        )

        self.assertIn("Type series", html)
        self.assertIn("Cecha wariantu", html)
        self.assertIn("type_series[].thickness.value", html)
        self.assertIn("Reguła wiersza i hierarchii", html)

    def test_web_ui_does_not_show_unaccepted_initial_product_model(self):
        html = render_home({"target_fields": [{"key": "product.name.value", "label": "Nazwa", "group": "Product identity"}]})

        self.assertIn("const INITIAL_PRODUCT_MODEL = {", html)
        self.assertIn("Wczytaj i zaakceptuj wymagane pliki modelu PIM.", html)
        self.assertNotIn('<div class="structure-field-label">Nazwa</div>', html)

    def test_maps_common_polish_product_columns(self):
        mapped = map_source_row(
            {
                "Nazwa produktu": "FAST TEST",
                "Kod produktu": "SKU-1",
                "Kategoria": "Tynki silikonowe",
                "Jednostka": "kg",
                "Opis": "Opis techniczny",
                "Lambda": "0,035",
            }
        )

        self.assertEqual(mapped["core"]["product_name"], "FAST TEST")
        self.assertEqual(mapped["core"]["external_id"], "SKU-1")
        self.assertEqual(mapped["core"]["unit"], "kg")
        self.assertEqual(mapped["product_information"]["description"], "Opis techniczny")
        self.assertEqual(mapped["sot"]["lambda"], "0,035")

    def test_converts_json_to_products_file(self):
        payload = json.dumps(
            [
                {
                    "Nazwa produktu": "FAST TEST",
                    "Kod produktu": "SKU-1",
                    "Kategoria": "Tynki silikonowe",
                    "Jednostka": "kg",
                    "Opis": "Opis techniczny",
                }
            ]
        ).encode("utf-8")

        with tempfile.TemporaryDirectory() as tmp:
            result = convert_uploaded_file("products.json", payload, Path(tmp))
            products_path = Path(tmp) / result["job_id"] / "products.json"
            data = json.loads(products_path.read_text(encoding="utf-8"))

        self.assertEqual(result["products_count"], 1)
        self.assertEqual(data["productsCount"], 1)
        product_attrs = data["products"][0]["dataVersions"][0]["productAttributes"]
        self.assertTrue(any(attr["AttributeId"] == 225 and attr["varcharValue"] == "FAST TEST" for attr in product_attrs))
        self.assertTrue(any(attr["AttributeId"] == 230 and attr["IntValue"] == 260 for attr in product_attrs))

    def test_converts_color_rows_to_colors_json_with_file_references(self):
        payload = json.dumps(
            [
                {"Name": "E1-10", "Type": "Kolor prosty", "RGB": "#9b5d74", "Roughness": "mat"},
                {"Name": "FG 01", "Type": "Tekstura", "Main": "textures/fg01/basecolor.png", "Normal": "https://example.test/fg01-normal.png"},
            ]
        ).encode("utf-8")

        analysis = analyze_colors_file("colors.json", payload)
        self.assertEqual(analysis["fields"][0]["key"], "name")
        self.assertIn("Name", analysis["tables"][0]["columns"])

        with tempfile.TemporaryDirectory() as tmp:
            result = convert_colors_file(
                "colors.json",
                payload,
                Path(tmp),
                color_mapping={
                    "name": "Name",
                    "type": "Type",
                    "colorRGB": "RGB",
                    "roughness": "Roughness",
                    "MainTexture": "Main",
                    "normal_map": "Normal",
                },
                color_choice_mapping={
                    "type": {"Kolor prosty": "simple", "Tekstura": "advanced"},
                    "roughness": {"mat": "matte"},
                },
            )
            colors_path = Path(tmp) / result["job_id"] / "colors.json"
            data = json.loads(colors_path.read_text(encoding="utf-8"))

        self.assertEqual(data["Count"], 2)
        first_params = data["colors"][0]["dataVersions"][0]["parameters"]
        self.assertTrue(any(item["parameterName"] == "r" and item["IntValue"] == 155 for item in first_params))
        self.assertTrue(any(item["parameterName"] == "type" and item["TextValue"] == "simple" for item in first_params))
        self.assertTrue(any(item["parameterName"] == "roughness" and item["TextValue"] == "matte" for item in first_params))
        second_version = data["colors"][1]["dataVersions"][0]
        self.assertTrue(any(item["parameterName"] == "type" and item["TextValue"] == "advanced" for item in second_version["parameters"]))
        self.assertTrue(any(item["parameterName"] == "MainTexture" and item["fileName"] == "basecolor.png" for item in second_version["filesParameters"]))
        self.assertTrue(any(item["parameterName"] == "normal_map" and item["fileUrl"] == "https://example.test/fg01-normal.png" for item in second_version["filesParameters"]))

    def test_maps_system_rows(self):
        mapped = map_system_row(
            {
                "Nazwa systemu": "ETICS TEST",
                "Zastosowanie systemu": "Ocieplenie scian od zewnatrz",
                "Rodzaj izolacji": "Welna mineralna",
                "Wariant": "Tynk silikonowy",
                "Pozycja warstwy": 1,
                "Nazwa warstwy": "Zaprawa klejaca",
                "Kod produktu": "SKU-1",
                "Domyslny": "tak",
                "Ilosc": 4,
            }
        )

        self.assertEqual(mapped["system"]["system_name"], "ETICS TEST")
        self.assertEqual(mapped["variant"]["variant_name"], "Tynk silikonowy")
        self.assertEqual(mapped["layer"]["layer_name"], "Zaprawa klejaca")
        self.assertEqual(mapped["product"]["product_code"], "SKU-1")

    def test_converts_json_to_products_and_building_elements_files(self):
        payload = json.dumps(
            [
                {
                    "Nazwa systemu": "ETICS TEST",
                    "Zastosowanie systemu": "Ocieplenie scian od zewnatrz",
                    "Rodzaj izolacji": "Welna mineralna",
                    "Typ BIM": "Wall",
                    "Wariant": "Tynk silikonowy",
                    "Pozycja warstwy": 1,
                    "Nazwa warstwy": "Zaprawa klejaca",
                    "Nazwa produktu": "FAST TEST",
                    "Kod produktu": "SKU-1",
                    "Kategoria": "Zaprawy klejowe",
                    "Jednostka": "kg",
                    "Domyslny": "tak",
                    "Ilosc": 4,
                }
            ]
        ).encode("utf-8")

        with tempfile.TemporaryDirectory() as tmp:
            result = convert_uploaded_file("systems.json", payload, Path(tmp))
            products_path = Path(tmp) / result["job_id"] / "products.json"
            elements_path = Path(tmp) / result["job_id"] / "building_elements.json"
            products_data = json.loads(products_path.read_text(encoding="utf-8"))
            elements_data = json.loads(elements_path.read_text(encoding="utf-8"))

        product_id = products_data["products"][0]["Id"]
        element_attrs = elements_data["buildingElements"][0]["dataVersions"][0]["productAttributes"]

        self.assertEqual(result["products_count"], 1)
        self.assertEqual(result["building_elements_count"], 1)
        self.assertEqual(elements_data["buildingElementsCount"], 1)
        self.assertTrue(any(attr["AttributeId"] == 280 and attr["varcharValue"] == "ETICS TEST" for attr in element_attrs))
        self.assertTrue(any(attr["AttributeId"] == 290 and attr["IntValue"] == product_id for attr in element_attrs))

    def test_converts_products_and_systems_separately(self):
        products_payload = json.dumps(
            [
                {
                    "Nazwa produktu": "FAST TEST",
                    "Kod produktu": "SKU-1",
                    "Kategoria": "Zaprawy klejowe",
                    "Jednostka": "kg",
                }
            ]
        ).encode("utf-8")
        systems_payload = json.dumps(
            [
                {
                    "Nazwa systemu": "ETICS TEST",
                    "Wariant": "Tynk silikonowy",
                    "Pozycja warstwy": 1,
                    "Nazwa warstwy": "Zaprawa klejaca",
                    "Kod produktu": "SKU-1",
                    "Domyslny": "tak",
                    "Ilosc": 4,
                }
            ]
        ).encode("utf-8")

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            products_result = convert_products_file("products.json", products_payload, root)
            generated_products = root / products_result["job_id"] / "products.json"
            systems_result = convert_systems_file(
                "systems.json",
                systems_payload,
                generated_products.read_bytes(),
                root,
            )
            elements_path = root / systems_result["job_id"] / "building_elements.json"
            elements_data = json.loads(elements_path.read_text(encoding="utf-8"))

        self.assertEqual(products_result["products_count"], 1)
        self.assertEqual(systems_result["building_elements_count"], 1)
        self.assertEqual(elements_data["buildingElementsCount"], 1)

    def test_suggests_product_and_system_mappings(self):
        rows = [
            {
                "Nazwa systemu": "ETICS TEST",
                "Wariant": "Tynk silikonowy",
                "Nazwa warstwy": "Zaprawa klejaca",
                "Kod produktu": "SKU-1",
                "Nazwa produktu": "FAST TEST",
            }
        ]

        product = suggest_mapping(rows, PRODUCT_FIELDS)
        system = suggest_mapping(rows, SYSTEM_FIELDS)

        self.assertEqual(product["mapping"]["Nazwa produktu"], "product.name.value")
        self.assertEqual(product["mapping"]["Kod produktu"], "product.code.value")
        self.assertEqual(system["mapping"]["Nazwa systemu"], "system_name")
        self.assertEqual(system["mapping"]["Wariant"], "variant_name")
        self.assertEqual(system["mapping"]["Nazwa warstwy"], "layer_name")

    def test_mapping_quality_reports_missing_product_values(self):
        rows = [
            {"Nazwa produktu": "FAST A", "Lambda": "0,035"},
            {"Nazwa produktu": "FAST B", "Lambda": ""},
            {"Nazwa produktu": "FAST C", "Lambda": "0,035"},
        ]

        product = suggest_mapping(rows, PRODUCT_FIELDS)
        quality = product["field_quality"]["Lambda"]

        self.assertEqual(quality["target_field"], "type_series[].lambda_value.value")
        self.assertEqual(quality["filled_rows"], 2)
        self.assertEqual(quality["missing_rows"], 1)
        self.assertEqual(quality["missing_row_numbers"], [2])
        self.assertEqual(quality["typical_values"][0]["value"], "0,035")

    def test_analysis_returns_column_samples_for_value_review(self):
        payload = json.dumps(
            {
                "products": [
                    {"name": "A", "parent_group": [22970]},
                    {"name": "B", "parent_group": [22815]},
                    {"name": "C", "parent_group": [22823]},
                ]
            }
        ).encode("utf-8")

        analysis = analyze_uploaded_file("products.json", payload)
        samples = analysis["tables"][0]["column_samples"]["parent_group"]

        self.assertEqual(samples[0]["row"], 1)
        self.assertEqual(samples[0]["value"], [22970])
        self.assertEqual(samples[2]["value"], [22823])

    def test_analysis_returns_more_preview_rows_for_shared_tables(self):
        payload = json.dumps(
            {
                "products": [
                    {"name": f"Product {index}", "lambda": "0,035"}
                    for index in range(75)
                ]
            }
        ).encode("utf-8")

        analysis = analyze_uploaded_file("products.json", payload)

        self.assertEqual(len(analysis["tables"][0]["sample_rows"]), 75)

    def test_json_label_value_cells_are_flattened_for_mapping(self):
        payload = json.dumps(
            {
                "products": [
                    {
                        "headline": {"label": "Nazwa produktu", "value": "Caparol Test"},
                        "text": {"label": "Krótki opis produktu", "value": "Opis"},
                    }
                ]
            }
        ).encode("utf-8")

        table = read_source_tables("products.json", payload)[0]

        self.assertEqual(table.rows[0]["headline"], "Caparol Test")
        self.assertEqual(table.rows[0]["text"], "Opis")

    def test_json_files_create_packshot_helper_columns(self):
        payload = json.dumps(
            {
                "products": [
                    {
                        "name": "Caparol Test",
                        "files": [
                            {
                                "type": "image",
                                "caption": "packshot.jpg",
                                "url": "https://example.com/assets/packshot.png",
                            },
                            {
                                "type": "ti",
                                "caption": "Karta techniczna",
                                "url": "https://example.com/assets/ti.pdf",
                            },
                        ],
                    }
                ]
            }
        ).encode("utf-8")

        table = read_source_tables("products.json", payload)[0]

        self.assertEqual(table.rows[0]["packshot_url"], "https://example.com/assets/packshot.png")
        self.assertEqual(table.rows[0]["image_urls"], "https://example.com/assets/packshot.png")
        self.assertEqual(table.rows[0]["document_urls"], "https://example.com/assets/ti.pdf")

    def test_csv_reader_accepts_windows_encoded_semicolon_files(self):
        payload = "Nazwa;Opis\nPłyta testowa;Zażółć gęślą jaźń\n".encode("cp1250")

        table = read_source_tables("produkty.csv", payload)[0]

        self.assertEqual(table.rows[0]["Nazwa"], "Płyta testowa")
        self.assertEqual(table.rows[0]["Opis"], "Zażółć gęślą jaźń")

    def test_files_column_is_suggested_as_documents_url(self):
        product = suggest_mapping([{"name": "Caparol Test", "files": [{"type": "image", "url": "https://example.com/a.png"}]}], PRODUCT_FIELDS)

        self.assertEqual(product["mapping"]["files"], "product.documents[].url.value")

    def test_product_files_export_to_document_attributes(self):
        payload = json.dumps(
            {
                "products": [
                    {
                        "name": "Caparol Test",
                        "files": [
                            {
                                "type": "image",
                                "caption": "packshot.jpg",
                                "url": "https://example.com/assets/packshot.png",
                            }
                        ],
                    }
                ]
            }
        ).encode("utf-8")

        with tempfile.TemporaryDirectory() as tmp:
            result = convert_products_file("products.json", payload, Path(tmp))
            products_path = Path(tmp) / result["job_id"] / "products.json"
            data = json.loads(products_path.read_text(encoding="utf-8"))

        attrs = data["products"][0]["dataVersions"][0]["productAttributes"]

        self.assertTrue(any(attr["AttributeId"] == 273 and attr["varcharValue"] == "https://example.com/assets/packshot.png" for attr in attrs))
        self.assertTrue(any(attr["AttributeId"] == 272 and attr["varcharValue"] == "image" for attr in attrs))
        self.assertTrue(any(attr["AttributeId"] == 274 and attr["varcharValue"] == "jpg" for attr in attrs))

    def test_product_files_export_through_mapping_profile(self):
        payload = json.dumps(
            {
                "products": [
                    {
                        "name": "Caparol Test",
                        "files": [
                            {
                                "type": "image",
                                "caption": "packshot.jpg",
                                "url": "https://example.com/assets/packshot.png",
                            }
                        ],
                    }
                ]
            }
        ).encode("utf-8")

        with tempfile.TemporaryDirectory() as tmp:
            result = convert_products_file(
                "products.json",
                payload,
                Path(tmp),
                product_mapping_profile={
                    "name": {"target_path": "product.name.value", "cleanup": {"trim": True}},
                    "files": {"target_path": "product.documents[].url.value", "cleanup": {"trim": True}},
                    "packshot_url": {"target_path": "product.documents[].url.value", "cleanup": {"trim": True}},
                },
            )
            products_path = Path(tmp) / result["job_id"] / "products.json"
            data = json.loads(products_path.read_text(encoding="utf-8"))

        attrs = data["products"][0]["dataVersions"][0]["productAttributes"]
        urls = [attr["varcharValue"] for attr in attrs if attr["AttributeId"] == 273]

        self.assertEqual(urls, ["https://example.com/assets/packshot.png"])

    def test_products_conversion_writes_client_mapping_report_xlsx(self):
        payload = json.dumps([{"Name": "FAST A", "Category": "Zaprawy"}]).encode("utf-8")
        profile = {
            "Name": {
                "source_column": "Name",
                "target_path": "product.name.value",
                "target_label": "Nazwa produktu",
                "target_group": "Product identity",
                "target_value_kind": "free_text",
                "cleanup": {"trim": True},
            },
            "Category": {
                "source_column": "Category",
                "target_path": "product.category[].value",
                "target_label": "Kategoria",
                "target_group": "Product identity",
                "target_value_kind": "multi_choice",
                "choice_map": {"Zaprawy": "Zaprawy klejowe"},
                "cleanup": {"trim": True},
            },
        }

        with tempfile.TemporaryDirectory() as tmp:
            result = convert_products_file("products.json", payload, Path(tmp), product_mapping_profile=profile)
            report_path = Path(tmp) / result["job_id"] / "mapping_report.xlsx"
            acceptance_path = Path(tmp) / result["job_id"] / "products_acceptance.xlsx"
            workbook = load_workbook(report_path)
            acceptance_workbook = load_workbook(acceptance_path)

        self.assertIn("mapping_report_xlsx", result["files"])
        self.assertIn("products_acceptance_xlsx", result["files"])
        self.assertIn("Mapowanie", workbook.sheetnames)
        self.assertIn("Mapy opcji", workbook.sheetnames)
        rows = list(workbook["Mapowanie"].iter_rows(values_only=True))
        self.assertIn("target_path", rows[0])
        self.assertTrue(any(row[2] == "product.name.value" and row[3] == "Nazwa produktu" for row in rows[1:]))
        option_rows = list(workbook["Mapy opcji"].iter_rows(values_only=True))
        self.assertTrue(any(row[3] == "Zaprawy" and row[4] == "Zaprawy klejowe" for row in option_rows[1:]))
        self.assertIn("Produkty", acceptance_workbook.sheetnames)
        self.assertIn("Cechy produktów", acceptance_workbook.sheetnames)
        product_rows = list(acceptance_workbook["Produkty"].iter_rows(values_only=True))
        self.assertIn("status_akceptacji", product_rows[0])
        self.assertIn("uwagi_klienta", product_rows[0])
        self.assertTrue(any(row[2] == "FAST A" for row in product_rows[1:]))
        detail_rows = list(acceptance_workbook["Cechy produktów"].iter_rows(values_only=True))
        self.assertIn("poprawiona_wartość", detail_rows[0])

    def test_product_export_uses_dynamic_type_series_ids_from_pim_model(self):
        model_files = {
            "productsModels.json": json.dumps(
                {
                    "models": [
                        {"Id": 66, "Name": "SG One Tool Product", "modelType": "Product"},
                        {"Id": 900, "Name": "Typoszereg", "modelType": "ProductRows"},
                    ]
                }
            ).encode("utf-8"),
            "productsAttributes.json": json.dumps(
                {
                    "attributes": [
                        {"Id": 225, "ProductModelId": 66, "AttributeName": "Nazwa", "AttributeType": "Text"},
                        {
                            "Id": 135,
                            "ProductModelId": 66,
                            "AttributeName": "Typoszereg",
                            "AttributeType": "Table_Model",
                            "TargetModelId": 900,
                            "SOTFlag": True,
                        },
                        {"Id": 501, "ProductModelId": 900, "AttributeName": "Grubość", "AttributeType": "Number", "Unit": "mm"},
                        {"Id": 502, "ProductModelId": 900, "AttributeName": "λD", "AttributeType": "Number", "Unit": "W/mK"},
                    ]
                }
            ).encode("utf-8"),
        }
        payload = json.dumps(
            [
                {
                    "Name": "Produkt testowy",
                    "Thickness": "100",
                    "Lambda": "0.035",
                }
            ]
        ).encode("utf-8")

        with tempfile.TemporaryDirectory() as tmp:
            result = convert_products_file(
                "products.json",
                payload,
                Path(tmp),
                product_mapping_profile={
                    "Name": {"target_path": "product.name.value"},
                    "Thickness": {"target_path": "type_series[].thickness.value"},
                    "Lambda": {"target_path": "type_series[].lambda_value.value"},
                },
                product_model_files=model_files,
            )
            products_path = Path(tmp) / result["job_id"] / "products.json"
            data = json.loads(products_path.read_text(encoding="utf-8"))

        attrs = data["products"][0]["dataVersions"][0]["productAttributes"]

        self.assertTrue(any(attr["AttributeId"] == 501 and attr["ParentAttributeId"] == 135 for attr in attrs))
        self.assertTrue(any(attr["AttributeId"] == 502 and attr["ParentAttributeId"] == 135 for attr in attrs))
        self.assertFalse(any(attr["ParentAttributeId"] == 276 and attr["AttributeId"] in {277, 278} for attr in attrs))

    def test_product_export_uses_dynamic_product_attribute_ids_from_pim_model(self):
        model_files = {
            "productsModels.json": json.dumps(
                {
                    "models": [
                        {"Id": 41, "Name": "SG One Tool Product", "modelType": "Product"},
                        {"Id": 45, "Name": "Typoszereg", "modelType": "Attribute"},
                    ]
                }
            ).encode("utf-8"),
            "productsAttributes.json": json.dumps(
                {
                    "attributes": [
                        {"Id": 116, "ProductModelId": 41, "AttributeName": "Nazwa", "AttributeType": "VarChar"},
                        {"Id": 128, "ProductModelId": 41, "AttributeName": "Opis", "AttributeType": "Longtext"},
                        {"Id": 129, "ProductModelId": 41, "AttributeName": "Link do strony www", "AttributeType": "Text"},
                        {"Id": 318, "ProductModelId": 41, "AttributeName": "PIM ID", "AttributeType": "VarChar"},
                        {
                            "Id": 135,
                            "ProductModelId": 41,
                            "AttributeName": "Typoszereg",
                            "AttributeType": "Table_Model",
                            "TargetModelId": 45,
                            "SOTFlag": True,
                        },
                        {"Id": 138, "ProductModelId": 45, "AttributeName": "λD", "AttributeType": "Number", "Unit": "W/mK"},
                        {"Id": 155, "ProductModelId": 45, "AttributeName": "SAP ID", "AttributeType": "VarChar"},
                        {"Id": 319, "ProductModelId": 45, "AttributeName": "Nazwa", "AttributeType": "VarChar"},
                        {"Id": 321, "ProductModelId": 45, "AttributeName": "PIM ID", "AttributeType": "VarChar"},
                    ]
                }
            ).encode("utf-8"),
        }
        payload = json.dumps(
            [
                {
                    "Name": "Produkt A",
                    "description": "Opis produktu",
                    "url": "https://example.com/product-a",
                    "pim.attribute.318.value": "PR0001",
                    "_type_series_rows": [
                        {
                            "pim.attribute.319.value": "11620533",
                            "pim.attribute.155.value": "AR00233378",
                            "pim.attribute.321.value": "AR00233378",
                            "type_series[].lambda_value.value": "0.035",
                        }
                    ],
                }
            ]
        ).encode("utf-8")

        with tempfile.TemporaryDirectory() as tmp:
            result = convert_products_file(
                "products.json",
                payload,
                Path(tmp),
                product_model_files=model_files,
            )
            products_path = Path(tmp) / result["job_id"] / "products.json"
            data = json.loads(products_path.read_text(encoding="utf-8"))

        product = data["products"][0]
        attrs = product["dataVersions"][0]["productAttributes"]
        attr_ids = {attr["AttributeId"] for attr in attrs}
        parent_ids = {attr["ParentAttributeId"] for attr in attrs}

        self.assertEqual(product["ModelType"], 41)
        self.assertTrue(any(attr["AttributeId"] == 116 and attr["varcharValue"] == "Produkt A" for attr in attrs))
        self.assertTrue(any(attr["AttributeId"] == 128 and attr["TextValue"] == "Opis produktu" for attr in attrs))
        self.assertTrue(any(attr["AttributeId"] == 129 and attr["TextValue"] == "https://example.com/product-a" for attr in attrs))
        self.assertTrue(any(attr["AttributeId"] == 318 and attr["varcharValue"] == "PR0001" for attr in attrs))
        self.assertTrue(any(attr["AttributeId"] == 138 and attr["ParentAttributeId"] == 135 for attr in attrs))
        self.assertTrue(any(attr["AttributeId"] == 319 and attr["ParentAttributeId"] == 135 and attr["varcharValue"] == "11620533" for attr in attrs))
        self.assertTrue(any(attr["AttributeId"] == 155 and attr["ParentAttributeId"] == 135 and attr["varcharValue"] == "AR00233378" for attr in attrs))
        self.assertTrue(any(attr["AttributeId"] == 321 and attr["ParentAttributeId"] == 135 and attr["varcharValue"] == "AR00233378" for attr in attrs))
        self.assertFalse({225, 226, 228, 229, 246, 303} & attr_ids)
        self.assertNotIn(233, parent_ids)

    def test_category_ids_accept_numeric_lists_and_json_arrays(self):
        self.assertEqual(category_ids([22970, 22815, 0]), [22815, 22970])
        self.assertEqual(category_ids("[22970, 22815]"), [22815, 22970])

    def test_parent_group_is_suggested_as_product_category(self):
        product = suggest_mapping([{"name": "Caparol Test", "parent_group": [22970]}], PRODUCT_FIELDS)

        self.assertEqual(product["mapping"]["parent_group"], "product.category[].value")

    def test_unknown_product_column_is_ignored_by_default(self):
        product = suggest_mapping([{"Opis marketingowy": "tekst"}], PRODUCT_FIELDS)

        self.assertEqual(product["mapping"]["Opis marketingowy"], "ignore")

    def test_product_mapping_targets_type_series_structure(self):
        rows = [
            {
                "Nazwa produktu": "FAST A",
                "Grubosc": "0,05",
                "Lambda": "0,035",
                "Gestosc": "100",
                "Cieplo wlasciwe": "1030",
            }
        ]

        product = suggest_mapping(rows, PRODUCT_FIELDS)

        self.assertEqual(product["mapping"]["Grubosc"], "type_series[].thickness.value")
        self.assertEqual(product["mapping"]["Lambda"], "type_series[].lambda_value.value")
        self.assertEqual(product["mapping"]["Gestosc"], "type_series[].density.value")
        self.assertEqual(product["mapping"]["Cieplo wlasciwe"], "type_series[].specific_heat.value")

    def test_applies_explicit_column_mapping_before_conversion(self):
        row = {
            "Towar": "FAST TEST",
            "Symbol": "SKU-1",
            "Opis klienta": "Opis techniczny",
        }
        mapped = apply_column_mapping(
            row,
            {
                "Towar": "product.name.value",
                "Symbol": "product.code.value",
                "Opis klienta": "product.description.value",
            },
        )

        self.assertEqual(mapped["product.name.value"], "FAST TEST")
        self.assertEqual(mapped["product.code.value"], "SKU-1")
        self.assertEqual(mapped["product.description.value"], "Opis techniczny")

    def test_applies_mapping_profile_cleanup_preview_rules(self):
        row = {
            "Grubosc opisowa": " ok. 100 mm ",
            "Jednostka": "mm",
        }
        mapped = apply_column_mapping_profile(
            row,
            {
                "Grubosc opisowa": {
                    "target_path": "type_series[].thickness.value",
                    "cleanup": {
                        "trim": True,
                        "removeText": "ok.",
                        "parseNumber": True,
                        "unitSourceColumn": "Jednostka",
                    },
                },
                "Jednostka": {"target_path": "ignore"},
            },
        )

        self.assertEqual(mapped["type_series[].thickness.value"], "100 mm")

    def test_mapping_profile_converts_source_unit_to_model_unit(self):
        mapped = apply_column_mapping_profile(
            {"Grubość": "100 mm"},
            {
                "Grubość": {
                    "target_path": "type_series[].thickness.value",
                    "cleanup": {
                        "parseNumber": True,
                        "unit": "mm",
                        "targetUnit": "m",
                        "unitConversionFactor": "0.001",
                    },
                }
            },
        )

        self.assertEqual(mapped["type_series[].thickness.value"], "0.1 m")

    def test_mapping_profile_extracts_two_cleaned_values_from_one_column(self):
        row = {
            "Wymiary": "gr. 12,5 mm / szer. 1200 mm",
        }

        mapped = apply_column_mapping_profile(
            row,
            {
                "Wymiary": {
                    "source_column": "Wymiary",
                    "target_path": "type_series[].thickness.value",
                    "cleanup": {
                        "splitBy": "/",
                        "splitPart": "1",
                        "removeText": "gr.",
                        "decimalComma": True,
                        "parseNumber": True,
                        "unit": "mm",
                    },
                },
                "Wymiary::extract::1": {
                    "source_column": "Wymiary",
                    "target_path": "pim.attribute.600.value",
                    "cleanup": {
                        "splitBy": "/",
                        "splitPart": "2",
                        "removeText": "szer.",
                        "parseNumber": True,
                        "unit": "mm",
                    },
                },
            },
        )

        self.assertEqual(mapped["type_series[].thickness.value"], "12.5 mm")
        self.assertEqual(mapped["pim.attribute.600.value"], "1200 mm")

    def test_unmapped_column_is_not_converted_to_custom_attribute(self):
        mapped = map_source_row(
            {
                "Nazwa produktu": "FAST TEST",
                "custom_attribute::Parametry ogniowe::Odporność ogniowa": "A1",
            }
        )

        self.assertNotIn("custom_attributes", mapped)
        self.assertEqual(mapped["unmapped"]["custom_attribute::Parametry ogniowe::Odporność ogniowa"], "A1")

    def test_custom_attribute_target_is_ignored(self):
        mapped = apply_column_mapping_profile(
            {"Cecha klienta": "A1"},
            {
                "Cecha klienta": {
                    "target_path": "custom_attribute",
                    "cleanup": {},
                }
            },
        )

        self.assertEqual(mapped, {})

    def test_mapping_profile_uses_row_type_context_for_groups(self):
        rows = [
            {"Typ": "grupa", "ID": "GR-1", "Parent ID": "", "Grupa": "Izolacje fasadowe"},
            {"Typ": "produkt", "ID": "FAST EPS 70", "Parent ID": "GR-1", "Kod": "EPS-70"},
        ]
        mapped = apply_mapping_profile_to_rows(
            rows,
            {
                "_row_rules": {
                    "row_type_column": "Typ",
                    "group_row_values": "grupa",
                    "product_row_values": "produkt",
                    "id_column": "ID",
                    "parent_id_column": "Parent ID",
                    "group_source_column": "Grupa",
                    "group_target_path": "product.category[].value",
                },
                "Typ": {"target_path": "ignore"},
                "ID": {"target_path": "product.name.value"},
                "Kod": {"target_path": "product.code.value"},
            },
        )

        self.assertEqual(len(mapped), 1)
        self.assertEqual(mapped[0]["product.name.value"], "FAST EPS 70")
        self.assertEqual(mapped[0]["product.code.value"], "EPS-70")
        self.assertEqual(mapped[0]["product.category[].value"], "Izolacje fasadowe")

    def test_mapping_profile_accepts_multiple_row_type_values(self):
        rows = [
            {"Typ": "produkt", "ID": "P-1", "Parent ID": "", "Nazwa": "System 1"},
            {"Typ": "wyrób", "ID": "P-2", "Parent ID": "", "Nazwa": "System 2"},
            {"Typ": "wariant", "ID": "V-1", "Parent ID": "P-1", "Nazwa": "Wariant 1"},
        ]
        mapped = apply_mapping_profile_to_rows(
            rows,
            {
                "_row_rules": {
                    "row_type_column": "Typ",
                    "product_row_values": "produkt; wyrób",
                    "group_row_values": "wariant",
                    "id_column": "ID",
                    "parent_id_column": "Parent ID",
                },
                "Nazwa": {"target_path": "product.name.value"},
            },
        )

        self.assertEqual([item["product.name.value"] for item in mapped], ["System 1", "System 2"])

    def test_mapping_profile_sets_database_product_type_id(self):
        rows = [
            {"Nazwa": "Produkt firmowy", "Typ": "1"},
            {"Nazwa": "Produkt obcy", "Typ": "3"},
        ]
        mapped = apply_mapping_profile_to_rows(
            rows,
            {
                "_product_type_rule": {
                    "source_column": "Typ",
                    "own_values": "1",
                    "other_values": "3",
                    "own_type_id": "1",
                    "other_type_id": "2",
                    "default_type_id": "1",
                },
                "Nazwa": {"target_path": "product.name.value"},
            },
        )
        products = [
            build_pim_product(map_source_row(row), index)
            for index, row in enumerate(mapped, start=1)
        ]

        self.assertEqual([product["prodctTypeId"] for product in products], [1, 2])

    def test_mapping_profile_uses_row_type_and_parent_id_for_group_context(self):
        rows = [
            {"Typ": "Product", "ID": "PR00064177", "Parent ID": "", "Nazwa grupy": "Izolacje fasadowe", "Nazwa produktu": ""},
            {"Typ": "Article", "ID": "AR00233844", "Parent ID": "PR00064177", "Nazwa grupy": "", "Nazwa produktu": "FAST EPS 70"},
        ]
        mapped = apply_mapping_profile_to_rows(
            rows,
            {
                "_row_rules": {
                    "row_type_column": "Typ",
                    "group_row_values": "Product",
                    "product_row_values": "Article",
                    "id_column": "ID",
                    "parent_id_column": "Parent ID",
                    "product_id_column": "ID",
                    "product_name_column": "Nazwa produktu",
                    "group_source_column": "Nazwa grupy",
                    "group_target_path": "product.category[].value",
                },
            },
        )

        self.assertEqual(len(mapped), 1)
        self.assertEqual(mapped[0]["product.code.value"], "AR00233844")
        self.assertEqual(mapped[0]["product.name.value"], "FAST EPS 70")
        self.assertEqual(mapped[0]["product.category[].value"], "Izolacje fasadowe")

    def test_mapping_profile_uses_object_type_and_parent_id_for_articles(self):
        rows = [
            {
                "<ID>": "AR00233844",
                "<Name>": "11620126",
                "<Parent ID>": "PR00058878",
                "<Object Type Name>": "Article",
                "Nazwa artykułu": "11620126",
            },
            {
                "<ID>": "PR00058878",
                "<Name>": "Rigips PRO typ H2 (GKBI) 1200x12,5",
                "<Parent ID>": "",
                "<Object Type Name>": "Product",
                "Nazwa Produktu": "Rigips PRO typ H2 (GKBI) 1200x12,5",
            },
        ]
        mapped = apply_mapping_profile_to_rows(
            rows,
            {
                "_row_rules": {
                    "row_type_column": "<Object Type Name>",
                    "group_row_values": "Product",
                    "product_row_values": "Article",
                    "id_column": "<ID>",
                    "parent_id_column": "<Parent ID>",
                    "product_id_column": "<ID>",
                    "product_name_column": "Nazwa artykułu",
                    "group_source_column": "Nazwa Produktu",
                    "group_target_path": "product.category[].value",
                },
                "<Object Type Name>": {"target_path": "ignore"},
                "<Parent ID>": {"target_path": "ignore"},
            },
        )

        self.assertEqual(len(mapped), 1)
        self.assertEqual(mapped[0]["product.code.value"], "AR00233844")
        self.assertEqual(mapped[0]["product.name.value"], "11620126")
        self.assertEqual(mapped[0]["product.category[].value"], "Rigips PRO typ H2 (GKBI) 1200x12,5")

    def test_row_rule_can_use_same_name_column_for_product_and_group(self):
        rows = [
            {
                "<ID>": "PR00064177",
                "<Name>": "RIGIPS 4PRO Fire+ typ DF 1200x12,5",
                "<Parent ID>": "",
                "<Object Type Name>": "Product",
            },
            {
                "<ID>": "AR00278759",
                "<Name>": "11620970",
                "<Parent ID>": "PR00064177",
                "<Object Type Name>": "Article",
            },
        ]
        mapped = apply_mapping_profile_to_rows(
            rows,
            {
                "_row_rules": {
                    "row_type_column": "<Object Type Name>",
                    "group_row_values": "Product",
                    "product_row_values": "Article",
                    "id_column": "<ID>",
                    "parent_id_column": "<Parent ID>",
                    "product_id_column": "<ID>",
                    "product_name_column": "<Name>",
                    "group_source_column": "<Name>",
                    "group_target_path": "product.category[].value",
                },
            },
        )

        self.assertEqual(len(mapped), 1)
        self.assertEqual(mapped[0]["product.code.value"], "AR00278759")
        self.assertEqual(mapped[0]["product.name.value"], "11620970")
        self.assertEqual(mapped[0]["product.category[].value"], "RIGIPS 4PRO Fire+ typ DF 1200x12,5")

    def test_product_variant_row_rule_builds_type_series_rows(self):
        rows = [
            {
                "<Name>": "Rigips PRO typ H2 (GKBI) 1200x12,5",
                "<Object Type Name>": "Product",
            },
            {
                "<Name>": "11620126",
                "<Object Type Name>": "Article",
                "Thickness": "12,5 mm",
            },
            {
                "<Name>": "11620122",
                "<Object Type Name>": "Article",
                "Thickness": "15 mm",
            },
        ]
        mapped = apply_mapping_profile_to_rows(
            rows,
            {
                "_row_rules": {
                    "row_mode": "product_variants",
                    "row_type_column": "<Object Type Name>",
                    "product_row_values": "Product",
                    "group_row_values": "Article",
                },
                "<Name>": {
                    "target_path": "product.name.value",
                },
                "<Name>::extract::1": {
                    "source_column": "<Name>",
                    "target_path": "type_series[].variant_name.value",
                },
                "<Name>::extract::2": {
                    "source_column": "<Name>",
                    "target_path": "pim.attribute.319.value",
                    "target_group": "Type series",
                },
                "Thickness": {
                    "target_path": "type_series[].thickness.value",
                    "cleanup": {"parseNumber": True, "decimalComma": True},
                },
            },
        )

        self.assertEqual(len(mapped), 1)
        self.assertEqual(mapped[0]["product.name.value"], "Rigips PRO typ H2 (GKBI) 1200x12,5")
        self.assertEqual(len(mapped[0]["_type_series_rows"]), 2)
        self.assertEqual(mapped[0]["_type_series_rows"][0]["type_series[].variant_code.value"], "1")
        self.assertEqual(mapped[0]["_type_series_rows"][0]["type_series[].variant_name.value"], "11620126")
        self.assertEqual(mapped[0]["_type_series_rows"][0]["pim.attribute.319.value"], "11620126")
        self.assertEqual(mapped[0]["_type_series_rows"][0]["type_series[].thickness.value"], "12.5")

        product = build_pim_product(map_source_row(mapped[0]), 1)
        thickness_attrs = [
            attr
            for attr in product["dataVersions"][0]["productAttributes"]
            if attr["AttributeId"] == 277
        ]
        self.assertEqual([attr["NumberValue"] for attr in thickness_attrs], [12.5, 15.0])
        self.assertEqual([attr["RowI"] for attr in thickness_attrs], [0, 1])

    def test_product_variant_row_rule_can_number_variants_without_variant_name(self):
        rows = [
            {"Type": "Product", "Name": "Board A"},
            {"Type": "Article", "Thickness": "10"},
            {"Type": "Article", "Thickness": "12"},
        ]
        mapped = apply_mapping_profile_to_rows(
            rows,
            {
                "_row_rules": {
                    "row_mode": "product_variants",
                    "row_type_column": "Type",
                    "product_row_values": "Product",
                    "group_row_values": "Article",
                },
                "Name": {"target_path": "product.name.value"},
                "Thickness": {"target_path": "type_series[].thickness.value"},
            },
        )

        self.assertEqual(mapped[0]["product.name.value"], "Board A")
        self.assertEqual(
            [row["type_series[].variant_code.value"] for row in mapped[0]["_type_series_rows"]],
            ["1", "2"],
        )
        self.assertNotIn("type_series[].variant_name.value", mapped[0]["_type_series_rows"][0])

    def test_mapping_profile_applies_multiple_row_hierarchy_rules(self):
        rows = [
            {"Typ": "Category", "ID": "CAT-1", "Nazwa": "Plyty gipsowe"},
            {"Typ": "BrandGroup", "ID": "BR-1", "Nazwa": "RIGIPS"},
            {"Typ": "Article", "ID": "AR-1", "ParentCategory": "CAT-1", "ParentBrand": "BR-1", "Nazwa": "11620126"},
        ]
        mapped = apply_mapping_profile_to_rows(
            rows,
            {
                "_row_rules": {
                    "rules": [
                        {
                            "row_type_column": "Typ",
                            "group_row_values": "Category",
                            "product_row_values": "Article",
                            "id_column": "ID",
                            "parent_id_column": "ParentCategory",
                            "group_source_column": "Nazwa",
                            "group_target_path": "product.category[].value",
                        },
                        {
                            "row_type_column": "Typ",
                            "group_row_values": "BrandGroup",
                            "product_row_values": "Article",
                            "id_column": "ID",
                            "parent_id_column": "ParentBrand",
                            "group_source_column": "Nazwa",
                            "group_target_path": "product.manufacturer.value",
                        },
                    ]
                },
                "ID": {"target_path": "product.code.value"},
                "Nazwa": {"target_path": "product.name.value"},
            },
        )

        self.assertEqual(len(mapped), 1)
        self.assertEqual(mapped[0]["product.code.value"], "AR-1")
        self.assertEqual(mapped[0]["product.name.value"], "11620126")
        self.assertEqual(mapped[0]["product.category[].value"], "Plyty gipsowe")
        self.assertEqual(mapped[0]["product.manufacturer.value"], "RIGIPS")

    def test_analyzes_uploaded_file_before_conversion(self):
        payload = json.dumps(
            [
                {
                    "Towar": "FAST TEST",
                    "Symbol": "SKU-1",
                    "Opis klienta": "Opis techniczny",
                }
            ]
        ).encode("utf-8")

        analysis = analyze_uploaded_file("client.json", payload)

        self.assertEqual(len(analysis["tables"]), 1)
        self.assertIn("Towar", analysis["tables"][0]["columns"])
        self.assertIn("product_mapping", analysis["tables"][0])
        self.assertIn("system_mapping", analysis["tables"][0])

    def test_analyzes_uploaded_file_with_product_model_json(self):
        payload = json.dumps([{"Odporność ogniowa": "A1"}]).encode("utf-8")
        model = json.dumps(
            {
                "target_fields": [
                    {
                        "key": "pim.attribute.501.value",
                        "label": "Odporność ogniowa",
                        "group": {"name": "Parametry ogniowe"},
                        "aliases": ["odpornosc ogniowa"],
                    }
                ]
            }
        ).encode("utf-8")

        analysis = analyze_uploaded_file("client.json", payload, product_model_content=model)
        product_mapping = analysis["tables"][0]["product_mapping"]

        self.assertEqual(product_mapping["mapping"]["Odporność ogniowa"], "pim.attribute.501.value")
        self.assertEqual(product_mapping["target_fields"][0]["group"], "Parametry ogniowe")

    def test_product_fields_from_pim_bundle_uses_models_and_attributes(self):
        files = {
            "productsModels.json": json.dumps(
                {
                    "models": [
                        {"Id": 66, "Name": "Product Model", "modelType": "Product"},
                        {"Id": 73, "Name": "SOT Model", "modelType": "Attribute"},
                    ]
                }
            ).encode("utf-8"),
            "productsAttributes.json": json.dumps(
                {
                    "attributes": [
                        {
                            "Id": 225,
                            "ProductModelId": 66,
                            "AttributeName": "product_name",
                            "DispName": "Name",
                            "AttributeType": "VarChar",
                            "DisplayOrder": 1,
                        },
                        {
                            "Id": 276,
                            "ProductModelId": 66,
                            "AttributeName": "sot_table",
                            "DispName": "Series of Types",
                            "AttributeType": "Table_Model",
                            "TargetModelId": 73,
                            "DisplayOrder": 2,
                        },
                        {
                            "Id": 277,
                            "ProductModelId": 73,
                            "AttributeName": "sot_thickness",
                            "DispName": "Thickness",
                            "AttributeType": "Number",
                            "Unit": "m",
                            "DisplayOrder": 1,
                        },
                        {
                            "Id": 296,
                            "ProductModelId": 66,
                            "AttributeName": "structure",
                            "DispName": "Colors",
                            "AttributeType": "Colors",
                            "DisplayOrder": 3,
                        },
                    ]
                }
            ).encode("utf-8"),
        }

        fields = product_fields_from_pim_bundle(files)
        keys = [field.key for field in fields]

        self.assertIn("product.name.value", keys)
        self.assertIn("type_series[].thickness.value", keys)
        self.assertNotIn("pim.attribute.296.value", keys)
        sot_field = next(field for field in fields if field.key == "type_series[].thickness.value")
        self.assertEqual(sot_field.group, "Type series")
        self.assertEqual(sot_field.unit, "m")

    def test_product_fields_from_pim_bundle_keeps_choice_options(self):
        files = {
            "productsModels.json": json.dumps(
                {"models": [{"Id": 41, "Name": "SG_One_Tool_Product", "modelType": "Product"}]}
            ).encode("utf-8"),
            "productsAttributes.json": json.dumps(
                {
                    "attributes": [
                        {
                            "Id": 130,
                            "ProductModelId": 41,
                            "AttributeName": "Funkcja",
                            "DispName": "Funkcja",
                            "AttributeType": "Checkboxes",
                            "AttributeOptions": [
                                {"Id": 162, "OptionName": "Izolacja termiczna", "OptionValue": "1"},
                                {"Id": 164, "OptionName": "Szkło", "OptionValue": "3"},
                            ],
                        },
                        {
                            "Id": 148,
                            "ProductModelId": 41,
                            "AttributeName": "Klasa reakcji na ogień",
                            "DispName": "Klasa reakcji na ogień",
                            "AttributeType": "Radio",
                            "AttributeOptions": [
                                {"Id": 173, "OptionName": "A1", "OptionValue": ""},
                            ],
                        },
                    ]
                }
            ).encode("utf-8"),
        }

        fields = product_fields_from_pim_bundle(files)
        function = next(field for field in fields if field.label == "Funkcja")
        fire_class = next(field for field in fields if field.label == "Klasa reakcji na ogień")

        self.assertEqual(function.value_kind, "multi_choice")
        self.assertEqual([option["id"] for option in function.options], [162, 164])
        self.assertEqual(fire_class.value_kind, "single_choice")

    def test_product_fields_from_pim_bundle_accepts_singular_export_file_names(self):
        fields = product_fields_from_pim_bundle(
            {
                "ProductModels.json": json.dumps(
                    {"models": [{"Id": 41, "Name": "SG_One_Tool_Product", "modelType": "Product"}]}
                ).encode("utf-8"),
                "ProductAttributes.json": json.dumps(
                    {
                        "attributes": [
                            {
                                "Id": 116,
                                "ProductModelId": 41,
                                "AttributeName": "Nazwa",
                                "DispName": "Nazwa",
                                "AttributeType": "TextBox",
                            }
                        ]
                    }
                ).encode("utf-8"),
                "Product.json": b"[]",
            }
        )

        self.assertTrue(any(field.key == "product.name.value" for field in fields))

    def test_analyzes_product_model_files_without_customer_data(self):
        model = analyze_product_model_files(
            {
                "productsModels.json": json.dumps(
                    {"models": [{"Id": 41, "Name": "SG_One_Tool_Product", "modelType": "Product"}]}
                ).encode("utf-8"),
                "productsAttributes.json": json.dumps(
                    {
                        "attributes": [
                            {
                                "Id": 116,
                                "ProductModelId": 41,
                                "AttributeName": "Nazwa",
                                "DispName": "Nazwa",
                                "AttributeType": "TextBox",
                            }
                        ]
                    }
                ).encode("utf-8"),
                "products.json": b"[]",
            }
        )

        self.assertEqual(model["status"], "model_loaded")
        self.assertTrue(any(field["key"] == "product.name.value" for field in model["target_fields"]))

    def test_choice_field_quality_and_export_use_pim_option_ids(self):
        fields = [
            PRODUCT_FIELDS[0],
            product_fields_from_pim_bundle(
                {
                    "productsModels.json": json.dumps(
                        {"models": [{"Id": 41, "Name": "Product", "modelType": "Product"}]}
                    ).encode("utf-8"),
                    "productsAttributes.json": json.dumps(
                        {
                            "attributes": [
                                {
                                    "Id": 130,
                                    "ProductModelId": 41,
                                    "AttributeName": "Funkcja",
                                    "DispName": "Funkcja",
                                    "AttributeType": "Checkboxes",
                                    "AttributeOptions": [
                                        {"Id": 162, "OptionName": "Izolacja termiczna", "OptionValue": "1"},
                                        {"Id": 164, "OptionName": "Szkło", "OptionValue": "3"},
                                    ],
                                }
                            ]
                        }
                    ).encode("utf-8"),
                }
            )[0],
        ]
        mapping = {"Funkcja": "pim.attribute.130.value"}
        quality = field_quality([{"Funkcja": "Izolacja termiczna; Szkło; Brak"}], mapping, fields)
        mapped = apply_column_mapping_profile(
            {"Nazwa": "Produkt", "Funkcja": "Izolacja termiczna; Szkło"},
            {
                "Nazwa": {"target_path": "product.name.value"},
                "Funkcja": {
                    "target_path": "pim.attribute.130.value",
                    "target_value_kind": "multi_choice",
                    "target_options": list(fields[1].options),
                },
            },
        )
        product = build_pim_product(map_source_row(mapped), 1)
        attrs = [
            attr
            for attr in product["dataVersions"][0]["productAttributes"]
            if attr["AttributeId"] == 130
        ]

        self.assertEqual(quality["Funkcja"]["choice_quality"]["matched_values"], ["Izolacja termiczna", "Szkło"])
        self.assertEqual(quality["Funkcja"]["choice_quality"]["unmatched_values"], ["Brak"])
        self.assertEqual([attr["IntValue"] for attr in attrs], [162, 164])
        self.assertTrue(all(attr["BooleanValue"] for attr in attrs))

    def test_choice_mapping_can_translate_client_values_to_pim_options(self):
        mapped = apply_column_mapping_profile(
            {"Function": "thermal; glass"},
            {
                "Function": {
                    "target_path": "pim.attribute.130.value",
                    "target_value_kind": "multi_choice",
                    "target_options": [
                        {"id": 162, "label": "Thermal insulation", "value": "1"},
                        {"id": 164, "label": "Glass", "value": "3"},
                    ],
                    "choice_map": {
                        "thermal": "Thermal insulation",
                        "glass": "Glass",
                    },
                },
            },
        )
        product = build_pim_product(map_source_row({"product.name.value": "Product", **mapped}), 1)
        attrs = [
            attr
            for attr in product["dataVersions"][0]["productAttributes"]
            if attr["AttributeId"] == 130
        ]

        self.assertEqual([attr["IntValue"] for attr in attrs], [162, 164])

    def test_unmapped_choice_values_are_not_exported(self):
        mapped = apply_column_mapping_profile(
            {"Function": "thermal; unknown"},
            {
                "Function": {
                    "target_path": "pim.attribute.130.value",
                    "target_value_kind": "multi_choice",
                    "target_options": [
                        {"id": 162, "label": "Thermal insulation", "value": "1"},
                        {"id": 164, "label": "Glass", "value": "3"},
                    ],
                    "choice_map": {
                        "thermal": "Thermal insulation",
                    },
                },
            },
        )
        product = build_pim_product(map_source_row({"product.name.value": "Product", **mapped}), 1)
        attrs = [
            attr
            for attr in product["dataVersions"][0]["productAttributes"]
            if attr["AttributeId"] == 130
        ]

        self.assertEqual([attr["IntValue"] for attr in attrs], [162])
        self.assertTrue(all(attr["varcharValue"] is None and attr["TextValue"] is None for attr in attrs))

    def test_unmapped_single_choice_value_is_not_exported_as_text(self):
        mapped = apply_column_mapping_profile(
            {"Fire class": "A2"},
            {
                "Fire class": {
                    "target_path": "pim.attribute.148.value",
                    "target_value_kind": "single_choice",
                    "target_options": [
                        {"id": 173, "label": "A1", "value": "1"},
                    ],
                },
            },
        )
        product = build_pim_product(map_source_row({"product.name.value": "Product", **mapped}), 1)
        attrs = [
            attr
            for attr in product["dataVersions"][0]["productAttributes"]
            if attr["AttributeId"] == 148
        ]

        self.assertEqual(attrs, [])

    def test_product_fields_from_pim_bundle_detects_type_series_by_sot_flag(self):
        files = {
            "productsModels.json": json.dumps(
                {
                    "models": [
                        {"Id": 41, "Name": "SG_One_Tool_Product", "modelType": "Product"},
                        {"Id": 45, "Name": "Dowolna nazwa modelu", "modelType": "Attribute"},
                    ]
                }
            ).encode("utf-8"),
            "productsAttributes.json": json.dumps(
                {
                    "attributes": [
                        {
                            "Id": 116,
                            "ProductModelId": 41,
                            "AttributeName": "Nazwa",
                            "DispName": "Nazwa",
                            "AttributeType": "VarChar",
                        },
                        {
                            "Id": 135,
                            "ProductModelId": 41,
                            "AttributeName": "Dowolna tabela",
                            "DispName": "Dowolna tabela",
                            "AttributeType": "Table_Model",
                            "TargetModelId": 45,
                            "SOTFlag": True,
                        },
                        {
                            "Id": 133,
                            "ProductModelId": 45,
                            "AttributeName": "Grubość",
                            "DispName": "Grubość",
                            "AttributeType": "Number",
                            "Unit": "m",
                        },
                        {
                            "Id": 138,
                            "ProductModelId": 45,
                            "AttributeName": "λD",
                            "DispName": "λD",
                            "AttributeType": "Number",
                            "Unit": "W/mK",
                        },
                        {
                            "Id": 139,
                            "ProductModelId": 45,
                            "AttributeName": "Ciepło właściwe",
                            "DispName": "Ciepło właściwe",
                            "AttributeType": "Number",
                            "Unit": "J/kgK",
                        },
                        {
                            "Id": 155,
                            "ProductModelId": 45,
                            "AttributeName": "m2/paczka",
                            "DispName": "m2/paczka",
                            "AttributeType": "Number",
                        },
                    ]
                }
            ).encode("utf-8"),
        }

        fields = product_fields_from_pim_bundle(files)
        keys = [field.key for field in fields]
        groups_by_label = {field.label: field.group for field in fields}

        self.assertIn("type_series[].thickness.value", keys)
        self.assertIn("type_series[].lambda_value.value", keys)
        self.assertIn("type_series[].specific_heat.value", keys)
        self.assertIn("pim.attribute.155.value", keys)
        self.assertEqual(groups_by_label["m2/paczka"], "Type series")

    def test_product_fields_from_pim_bundle_detects_type_series_by_target_model_name(self):
        files = {
            "productsModels.json": json.dumps(
                {
                    "models": [
                        {"Id": 10, "Name": "Product Model", "modelType": "Product"},
                        {"Id": 20, "Name": "Typoszereg", "modelType": "Attribute"},
                    ]
                }
            ).encode("utf-8"),
            "productsAttributes.json": json.dumps(
                {
                    "attributes": [
                        {
                            "Id": 500,
                            "ProductModelId": 10,
                            "AttributeName": "Tabela parametrów",
                            "DispName": "Tabela parametrów",
                            "AttributeType": "Table_Model",
                            "TargetModelId": 20,
                        },
                        {
                            "Id": 501,
                            "ProductModelId": 20,
                            "AttributeName": "Parametr klienta",
                            "DispName": "Parametr klienta",
                            "AttributeType": "Number",
                        },
                    ]
                }
            ).encode("utf-8"),
        }

        fields = product_fields_from_pim_bundle(files)

        self.assertEqual(fields[0].key, "pim.attribute.501.value")
        self.assertEqual(fields[0].group, "Type series")

    def test_product_fields_from_pim_bundle_detects_technical_table_without_type_series_labels(self):
        files = {
            "productsModels.json": json.dumps(
                {
                    "models": [
                        {"Id": 10, "Name": "Product Model", "modelType": "Product"},
                        {"Id": 20, "Name": "Tabela parametrów", "modelType": "Attribute"},
                    ]
                }
            ).encode("utf-8"),
            "productsAttributes.json": json.dumps(
                {
                    "attributes": [
                        {
                            "Id": 500,
                            "ProductModelId": 10,
                            "AttributeName": "Parametry techniczne",
                            "DispName": "Parametry techniczne",
                            "AttributeType": "Table_Model",
                            "TargetModelId": 20,
                            "SOTFlag": False,
                        },
                        {
                            "Id": 501,
                            "ProductModelId": 20,
                            "AttributeName": "Parametr A",
                            "DispName": "Parametr A",
                            "AttributeType": "Number",
                            "Unit": "mm",
                        },
                        {
                            "Id": 502,
                            "ProductModelId": 20,
                            "AttributeName": "Parametr B",
                            "DispName": "Parametr B",
                            "AttributeType": "Number",
                            "Unit": "W/mK",
                        },
                    ]
                }
            ).encode("utf-8"),
        }

        fields = product_fields_from_pim_bundle(files)

        self.assertEqual([field.group for field in fields], ["Type series", "Type series"])
        self.assertEqual([field.key for field in fields], ["pim.attribute.501.value", "pim.attribute.502.value"])

    def test_product_fields_from_pim_bundle_does_not_treat_document_table_as_type_series(self):
        files = {
            "productsModels.json": json.dumps(
                {
                    "models": [
                        {"Id": 10, "Name": "Product Model", "modelType": "Product"},
                        {"Id": 20, "Name": "Documents Model", "modelType": "Attribute"},
                    ]
                }
            ).encode("utf-8"),
            "productsAttributes.json": json.dumps(
                {
                    "attributes": [
                        {
                            "Id": 500,
                            "ProductModelId": 10,
                            "AttributeName": "Documents",
                            "DispName": "Documents",
                            "AttributeType": "Table_Model",
                            "TargetModelId": 20,
                        },
                        {
                            "Id": 501,
                            "ProductModelId": 20,
                            "AttributeName": "File size",
                            "DispName": "File size",
                            "AttributeType": "Number",
                        },
                        {
                            "Id": 502,
                            "ProductModelId": 20,
                            "AttributeName": "File order",
                            "DispName": "File order",
                            "AttributeType": "Number",
                        },
                    ]
                }
            ).encode("utf-8"),
        }

        fields = product_fields_from_pim_bundle(files)

        self.assertNotIn("Type series", [field.group for field in fields])
        self.assertEqual(fields[0].group, "Documents")

    def test_product_fields_from_pim_bundle_ignores_deleted_attributes_even_when_used_by_products(self):
        files = {
            "productsModels.json": json.dumps(
                {
                    "models": [
                        {"Id": 10, "Name": "Product Model", "modelType": "Product"},
                    ]
                }
            ).encode("utf-8"),
            "productsAttributes.json": json.dumps(
                {
                    "attributes": [
                        {
                            "Id": 501,
                            "ProductModelId": 10,
                            "AttributeName": "Parametr archiwalny",
                            "DispName": "Parametr archiwalny",
                            "AttributeType": "Number",
                            "deleted": True,
                        },
                    ]
                }
            ).encode("utf-8"),
            "products.json": json.dumps(
                {
                    "products": [
                        {
                            "dataVersions": [
                                {
                                    "productAttributes": [
                                        {"AttributeId": 501, "ParentAttributeId": 0, "NumberValue": 10},
                                    ]
                                }
                            ]
                        }
                    ]
                }
            ).encode("utf-8"),
        }

        fields = product_fields_from_pim_bundle(files)

        self.assertEqual(fields, [])

    def test_product_fields_from_pim_bundle_ignores_access_model_attributes_used_by_products(self):
        files = {
            "productsModels.json": json.dumps(
                {
                    "models": [
                        {"Id": 10, "Name": "Product Model", "modelType": "Product"},
                        {"Id": 20, "Name": "Access Model", "modelType": "Access"},
                    ]
                }
            ).encode("utf-8"),
            "productsAttributes.json": json.dumps(
                {
                    "attributes": [
                        {
                            "Id": 501,
                            "ProductModelId": 10,
                            "AttributeName": "Nazwa",
                            "DispName": "Nazwa",
                            "AttributeType": "VarChar",
                        },
                        {
                            "Id": 777,
                            "ProductModelId": 20,
                            "AttributeName": "Grupa produktów",
                            "DispName": "Grupa produktów",
                            "AttributeType": "Checkboxes",
                        },
                    ]
                }
            ).encode("utf-8"),
            "products.json": json.dumps(
                {
                    "products": [
                        {
                            "dataVersions": [
                                {
                                    "productAttributes": [
                                        {"AttributeId": 501, "ParentAttributeId": 0, "varcharValue": "Produkt"},
                                        {"AttributeId": 777, "ParentAttributeId": 0, "IntValue": 1},
                                    ]
                                }
                            ]
                        }
                    ]
                }
            ).encode("utf-8"),
        }

        fields = product_fields_from_pim_bundle(files)
        keys = [field.key for field in fields]

        self.assertIn("product.name.value", keys)
        self.assertNotIn("pim.attribute.777.value", keys)

    def test_product_fields_from_pim_bundle_does_not_parse_products_data_file(self):
        files = {
            "productsModels.json": json.dumps(
                {
                    "models": [
                        {"Id": 66, "Name": "Product Model", "modelType": "Product"},
                    ]
                }
            ).encode("utf-8"),
            "productsAttributes.json": json.dumps(
                {
                    "attributes": [
                        {
                            "Id": 225,
                            "ProductModelId": 66,
                            "AttributeName": "product_name",
                            "DispName": "Name",
                            "AttributeType": "VarChar",
                        },
                    ]
                }
            ).encode("utf-8"),
            "products.json": b"{this file is deliberately not parsed as model metadata",
        }

        fields = product_fields_from_pim_bundle(files)

        self.assertEqual(fields[0].key, "product.name.value")

    def test_analyzes_uploaded_file_with_pim_model_bundle(self):
        payload = json.dumps([{"Thickness": "0,12"}]).encode("utf-8")
        files = {
            "productsModels.json": json.dumps(
                {
                    "models": [
                        {"Id": 66, "Name": "Product Model", "modelType": "Product"},
                        {"Id": 73, "Name": "SOT Model", "modelType": "Attribute"},
                    ]
                }
            ).encode("utf-8"),
            "productsAttributes.json": json.dumps(
                {
                    "attributes": [
                        {
                            "Id": 276,
                            "ProductModelId": 66,
                            "AttributeName": "sot_table",
                            "DispName": "Series of Types",
                            "AttributeType": "Table_Model",
                            "TargetModelId": 73,
                        },
                        {
                            "Id": 277,
                            "ProductModelId": 73,
                            "AttributeName": "sot_thickness",
                            "DispName": "Thickness",
                            "AttributeType": "Number",
                        },
                    ]
                }
            ).encode("utf-8"),
        }

        analysis = analyze_uploaded_file("client.json", payload, product_model_files=files)
        product_mapping = analysis["tables"][0]["product_mapping"]

        self.assertEqual(product_mapping["mapping"]["Thickness"], "type_series[].thickness.value")

    def test_custom_attribute_input_is_not_exported(self):
        payload = json.dumps(
            [
                {
                    "Nazwa produktu": "FAST TEST",
                    "custom_attribute::Parametry ogniowe::Odporność ogniowa": "A1",
                }
            ]
        ).encode("utf-8")

        with tempfile.TemporaryDirectory() as tmp:
            result = convert_uploaded_file("products.json", payload, Path(tmp))
            products_path = Path(tmp) / result["job_id"] / "products.json"
            data = json.loads(products_path.read_text(encoding="utf-8"))

        attrs = data["products"][0]["dataVersions"][0]["productAttributes"]

        self.assertFalse(any(attr["AttributeId"] in {238, 239, 240, 241, 242} for attr in attrs))

    def test_explicit_pim_attribute_mapping_exports_attribute_id(self):
        payload = json.dumps(
            [
                {
                    "Nazwa produktu": "FAST TEST",
                    "pim.attribute.243.value": "Model handlowy",
                }
            ]
        ).encode("utf-8")

        with tempfile.TemporaryDirectory() as tmp:
            result = convert_uploaded_file("products.json", payload, Path(tmp))
            products_path = Path(tmp) / result["job_id"] / "products.json"
            data = json.loads(products_path.read_text(encoding="utf-8"))

        attrs = data["products"][0]["dataVersions"][0]["productAttributes"]

        self.assertTrue(any(attr["AttributeId"] == 243 and attr["varcharValue"] == "Model handlowy" for attr in attrs))
        self.assertFalse(any(attr["AttributeId"] == 238 and attr["varcharValue"] == "pim.attribute.243.value" for attr in attrs))

    def test_manual_enrichment_fills_only_missing_values_by_default(self):
        rows = [
            {"product.name.value": "Produkt A"},
            {"product.name.value": "Produkt B", "pim.attribute.243.value": "Istniejąca wartość"},
        ]
        session = {
            "manual_entries": [
                {
                    "source": "manual",
                    "target_path": "pim.attribute.243.value",
                    "value": "Uzupełnienie",
                    "scope": "all",
                    "mode": "missing_only",
                }
            ],
            "typical_sources": [],
        }

        report = apply_enrichment_session_to_rows(rows, session)

        self.assertEqual(rows[0]["pim.attribute.243.value"], "Uzupełnienie")
        self.assertEqual(rows[1]["pim.attribute.243.value"], "Istniejąca wartość")
        self.assertEqual(report["applied_count"], 1)

    def test_manual_enrichment_updates_selected_type_series_row(self):
        rows = [
            {
                "product.name.value": "Produkt A",
                "_type_series_rows": [
                    {"type_series[].variant_name.value": "Wariant 1", "type_series[].lambda_value.value": "0.031"},
                    {"type_series[].variant_name.value": "Wariant 2", "type_series[].lambda_value.value": "0.032"},
                    {"type_series[].variant_name.value": "Wariant 3", "type_series[].lambda_value.value": "0.033"},
                ],
            }
        ]
        session = {
            "manual_entries": [
                {
                    "source": "manual",
                    "target_path": "type_series[].lambda_value.value",
                    "value": "0.040",
                    "scope": "current_product",
                    "product_key": "Produkt A",
                    "mode": "replace",
                    "apply_type_series_to_all": False,
                    "type_series_row_index": 1,
                }
            ],
            "typical_sources": [],
        }

        report = apply_enrichment_session_to_rows(rows, session)

        self.assertEqual(rows[0]["_type_series_rows"][0]["type_series[].lambda_value.value"], "0.031")
        self.assertEqual(rows[0]["_type_series_rows"][1]["type_series[].lambda_value.value"], "0.040")
        self.assertEqual(rows[0]["_type_series_rows"][2]["type_series[].lambda_value.value"], "0.033")
        self.assertEqual(report["applied_count"], 1)
        self.assertEqual(report["applied"][0]["type_series_row_index"], 1)

    def test_typical_products_fill_missing_pim_attributes_by_identity(self):
        products = [
            {
                "Id": 1,
                "dataVersions": [
                    {
                        "productAttributes": [
                            {"AttributeId": 225, "varcharValue": "Produkt A", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0}
                        ]
                    }
                ],
            }
        ]
        typical_payload = {
            "products": [
                {
                    "Id": 2,
                    "dataVersions": [
                        {
                            "productAttributes": [
                                {"AttributeId": 225, "varcharValue": "Produkt A", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0},
                                {"AttributeId": 243, "varcharValue": "Wartość typowa", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 238, "RowI": 0},
                            ]
                        }
                    ],
                }
            ]
        }

        report = apply_typical_products_to_products(products, typical_payload)
        attrs = products[0]["dataVersions"][0]["productAttributes"]

        self.assertTrue(any(attr["AttributeId"] == 243 and attr["varcharValue"] == "Wartość typowa" for attr in attrs))
        self.assertEqual(report["applied_count"], 1)

    def test_typical_products_can_be_matched_manually_when_names_differ(self):
        products = [
            {
                "Id": 1,
                "dataVersions": [
                    {
                        "productAttributes": [
                            {"AttributeId": 225, "varcharValue": "Produkt importowany", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0}
                        ]
                    }
                ],
            }
        ]
        typical_payload = {
            "products": [
                {
                    "Id": 2,
                    "dataVersions": [
                        {
                            "productAttributes": [
                                {"AttributeId": 225, "varcharValue": "Inna nazwa produktu typowego", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0},
                                {"AttributeId": 243, "varcharValue": "Dane typowe", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 238, "RowI": 0},
                            ]
                        }
                    ],
                }
            ]
        }
        session = {
            "typical_matches": [
                {
                    "product_key": "Produkt importowany",
                    "typical_key": "Inna nazwa produktu typowego",
                    "selected_attributes": ["243:238:0"],
                }
            ]
        }

        report = apply_typical_products_to_products(products, typical_payload, session)
        attrs = products[0]["dataVersions"][0]["productAttributes"]

        self.assertTrue(any(attr["AttributeId"] == 243 and attr["varcharValue"] == "Dane typowe" for attr in attrs))
        self.assertEqual(report["applied"][0]["match_source"], "manual_match")

    def test_typical_product_match_respects_selected_attributes(self):
        products = [
            {
                "Id": 1,
                "dataVersions": [
                    {
                        "productAttributes": [
                            {"AttributeId": 225, "varcharValue": "Produkt importowany", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0}
                        ]
                    }
                ],
            }
        ]
        typical_payload = {
            "products": [
                {
                    "Id": 2,
                    "dataVersions": [
                        {
                            "productAttributes": [
                                {"AttributeId": 225, "varcharValue": "Produkt typowy", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0},
                                {"AttributeId": 243, "varcharValue": "Nie kopiuj", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 238, "RowI": 0},
                                {"AttributeId": 244, "varcharValue": "Kopiuj", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 238, "RowI": 0},
                            ]
                        }
                    ],
                }
            ]
        }
        session = {
            "typical_matches": [
                {
                    "product_key": "Produkt importowany",
                    "typical_key": "Produkt typowy",
                    "selected_attributes": ["244:238:0"],
                }
            ]
        }

        apply_typical_products_to_products(products, typical_payload, session)
        attrs = products[0]["dataVersions"][0]["productAttributes"]

        self.assertFalse(any(attr["AttributeId"] == 243 for attr in attrs))
        self.assertTrue(any(attr["AttributeId"] == 244 and attr["varcharValue"] == "Kopiuj" for attr in attrs))

    def test_imported_product_can_use_multiple_typical_products(self):
        products = [
            {
                "Id": 1,
                "dataVersions": [
                    {
                        "productAttributes": [
                            {"AttributeId": 225, "varcharValue": "Produkt importowany", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0}
                        ]
                    }
                ],
            }
        ]
        typical_payload = {
            "products": [
                {
                    "Id": 2,
                    "dataVersions": [
                        {
                            "productAttributes": [
                                {"AttributeId": 225, "varcharValue": "Typowy A", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0},
                                {"AttributeId": 243, "varcharValue": "A", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 238, "RowI": 0},
                            ]
                        }
                    ],
                },
                {
                    "Id": 3,
                    "dataVersions": [
                        {
                            "productAttributes": [
                                {"AttributeId": 225, "varcharValue": "Typowy B", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0},
                                {"AttributeId": 244, "varcharValue": "B", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 238, "RowI": 0},
                            ]
                        }
                    ],
                },
            ]
        }
        session = {
            "typical_matches": [
                {"product_key": "Produkt importowany", "typical_key": "Typowy A", "selected_attributes": ["243:238:0"]},
                {"product_key": "Produkt importowany", "typical_key": "Typowy B", "selected_attributes": ["244:238:0"]},
            ]
        }

        report = apply_typical_products_to_products(products, typical_payload, session)
        attrs = products[0]["dataVersions"][0]["productAttributes"]

        self.assertTrue(any(attr["AttributeId"] == 243 and attr["varcharValue"] == "A" for attr in attrs))
        self.assertTrue(any(attr["AttributeId"] == 244 and attr["varcharValue"] == "B" for attr in attrs))
        self.assertEqual(report["applied_count"], 2)

    def test_manual_typical_match_with_empty_selection_copies_nothing(self):
        products = [
            {
                "Id": 1,
                "dataVersions": [
                    {
                        "productAttributes": [
                            {"AttributeId": 225, "varcharValue": "Produkt importowany", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0}
                        ]
                    }
                ],
            }
        ]
        typical_payload = {
            "products": [
                {
                    "Id": 2,
                    "dataVersions": [
                        {
                            "productAttributes": [
                                {"AttributeId": 225, "varcharValue": "Typowy A", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0},
                                {"AttributeId": 243, "varcharValue": "A", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 238, "RowI": 0},
                            ]
                        }
                    ],
                }
            ]
        }
        session = {
            "typical_matches": [
                {"source": "typical_manual_match", "product_key": "Produkt importowany", "typical_key": "Typowy A", "selected_attributes": []}
            ]
        }

        report = apply_typical_products_to_products(products, typical_payload, session)
        attrs = products[0]["dataVersions"][0]["productAttributes"]

        self.assertFalse(any(attr["AttributeId"] == 243 for attr in attrs))
        self.assertEqual(report["applied_count"], 0)

    def test_typical_enrichment_can_copy_selected_type_series_rows(self):
        products = [
            {
                "Id": 1,
                "dataVersions": [
                    {
                        "productAttributes": [
                            {"AttributeId": 225, "varcharValue": "Produkt importowany", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0}
                        ]
                    }
                ],
            }
        ]
        typical_payload = {
            "products": [
                {
                    "Id": 2,
                    "dataVersions": [
                        {
                            "productAttributes": [
                                {"AttributeId": 225, "varcharValue": "Produkt typowy", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0},
                                {"AttributeId": 243, "varcharValue": "Kopiuj", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 238, "RowI": 0},
                                {"AttributeId": 277, "varcharValue": "100", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 276, "RowI": 1},
                            ]
                        }
                    ],
                }
            ]
        }
        session = {
            "typical_matches": [
                {
                    "source": "typical_manual_match",
                    "product_key": "Produkt importowany",
                    "typical_key": "Produkt typowy",
                    "selected_attributes": ["243:238:0", "277:276:1"],
                }
            ]
        }

        report = apply_typical_products_to_products(products, typical_payload, session)
        attrs = products[0]["dataVersions"][0]["productAttributes"]

        self.assertTrue(any(attr["AttributeId"] == 243 and attr["varcharValue"] == "Kopiuj" for attr in attrs))
        self.assertTrue(any(attr["AttributeId"] == 277 and attr["ParentAttributeId"] == 276 and attr["RowI"] == 1 for attr in attrs))
        self.assertEqual(report["applied_count"], 2)

    def test_typical_enrichment_can_copy_type_series_to_all_variant_rows(self):
        products = [
            {
                "Id": 1,
                "dataVersions": [
                    {
                        "productAttributes": [
                            {"AttributeId": 225, "varcharValue": "Produkt importowany", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0},
                            {"AttributeId": 278, "varcharValue": "0.035", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 276, "RowI": 1},
                            {"AttributeId": 278, "varcharValue": "0.035", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 276, "RowI": 2},
                        ]
                    }
                ],
            }
        ]
        typical_payload = {
            "products": [
                {
                    "Id": 2,
                    "dataVersions": [
                        {
                            "productAttributes": [
                                {"AttributeId": 225, "varcharValue": "Produkt typowy", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0},
                                {"AttributeId": 277, "varcharValue": "100", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 276, "RowI": 1},
                            ]
                        }
                    ],
                }
            ]
        }
        session = {
            "typical_matches": [
                {
                    "source": "typical_manual_match",
                    "product_key": "Produkt importowany",
                    "typical_key": "Produkt typowy",
                    "selected_attributes": ["277:276:1"],
                    "apply_type_series_to_all": True,
                }
            ]
        }

        report = apply_typical_products_to_products(products, typical_payload, session)
        attrs = products[0]["dataVersions"][0]["productAttributes"]
        thickness_rows = sorted(attr["RowI"] for attr in attrs if attr["AttributeId"] == 277 and attr["ParentAttributeId"] == 276)

        self.assertEqual(thickness_rows, [1, 2])
        self.assertEqual(report["applied_count"], 2)

    def test_typical_enrichment_can_map_source_attribute_to_active_model_attribute(self):
        products = [
            {
                "Id": 1,
                "dataVersions": [
                    {
                        "productAttributes": [
                            {"AttributeId": 225, "varcharValue": "Produkt importowany", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0}
                        ]
                    }
                ],
            }
        ]
        typical_payload = {
            "products": [
                {
                    "Id": 2,
                    "dataVersions": [
                        {
                            "productAttributes": [
                                {"AttributeId": 116, "varcharValue": "Produkt typowy", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0},
                                {"AttributeId": 128, "varcharValue": None, "TextValue": "Opis ze starego modelu", "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0},
                            ]
                        }
                    ],
                }
            ]
        }
        session = {
            "typical_attribute_map": {
                "128:0": {"attribute_id": 246, "parent_attribute_id": 238}
            },
            "typical_matches": [
                {
                    "source": "typical_manual_match",
                    "product_key": "Produkt importowany",
                    "typical_key": "Produkt typowy",
                    "selected_attributes": ["128:0:0"],
                }
            ],
        }

        report = apply_typical_products_to_products(products, typical_payload, session)
        attrs = products[0]["dataVersions"][0]["productAttributes"]

        self.assertTrue(any(attr["AttributeId"] == 246 and attr["ParentAttributeId"] == 238 and attr["TextValue"] == "Opis ze starego modelu" for attr in attrs))
        self.assertFalse(any(attr["AttributeId"] == 128 for attr in attrs))
        self.assertEqual(report["applied_count"], 1)

    def test_typical_products_accept_uppercase_export_keys(self):
        products = [
            {
                "Id": 1,
                "DataVersions": [
                    {
                        "ProductAttributes": [
                            {"AttributeId": 225, "varcharValue": "Produkt importowany", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0}
                        ]
                    }
                ],
            }
        ]
        typical_payload = {
            "Products": [
                {
                    "Id": 2,
                    "ProductAttributes": [
                        {"AttributeId": 225, "varcharValue": "Typowy A", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 0, "RowI": 0},
                        {"AttributeId": 243, "varcharValue": "A", "TextValue": None, "NumberValue": None, "IntValue": None, "IntValue2": None, "BooleanValue": False, "ParentAttributeId": 238, "RowI": 0},
                    ],
                }
            ]
        }
        session = {
            "typical_matches": [
                {"source": "typical_manual_match", "product_key": "Produkt importowany", "typical_key": "Typowy A", "selected_attributes": ["243:238:0"]}
            ]
        }

        report = apply_typical_products_to_products(products, typical_payload, session)
        attrs = products[0]["DataVersions"][0]["ProductAttributes"]

        self.assertTrue(any(attr["AttributeId"] == 243 and attr["varcharValue"] == "A" for attr in attrs))
        self.assertEqual(report["applied_count"], 1)


if __name__ == "__main__":
    unittest.main()
