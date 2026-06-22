from __future__ import annotations

from io import BytesIO
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
TECH_FILL = PatternFill("solid", fgColor="E7E6E6")
WARNING_FILL = PatternFill("solid", fgColor="FCE4D6")
ACCEPT_FILL = PatternFill("solid", fgColor="D9EAF7")


def mapping_report_xlsx_bytes(report: dict[str, Any]) -> bytes:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Instrukcja"

    _write_rows(
        overview,
        [
            ["BuildData AI - raport mapowania", ""],
            ["Plik źródłowy", report.get("source_file", "")],
            ["Produkty wygenerowane", report.get("products_generated", "")],
            ["Elementy budowlane wygenerowane", report.get("building_elements_generated", "")],
            ["", ""],
            ["Jak używać", "Arkusz 'Mapowanie' zawiera docelowe cechy PIM i kolumny klienta. Klient może poprawić kolumny źródłowe, czyszczenie oraz mapę opcji."],
            ["Import zwrotny", "Nie usuwaj kolumn technicznych profile_key i target_path. Dzięki nim poprawiony Excel będzie możliwy do automatycznego wczytania."],
        ],
    )
    overview["A1"].font = Font(bold=True, size=14)

    _mapping_sheet(workbook, report)
    _choice_sheet(workbook, report)
    _row_rules_sheet(workbook, report)
    _tables_sheet(workbook, report)
    _unmapped_sheet(workbook, report)
    _warnings_sheet(workbook, report)
    _json_sheet(workbook, report)

    for sheet in workbook.worksheets:
        _format_sheet(sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def product_acceptance_xlsx_bytes(
    products_payload: dict[str, Any],
    *,
    source_file: str = "",
    attribute_labels: dict[int, str] | None = None,
) -> bytes:
    """Create a client-facing workbook for reviewing generated products."""
    attribute_labels = attribute_labels or {}
    products = products_payload.get("products") or []
    workbook = Workbook()
    instruction = workbook.active
    instruction.title = "Instrukcja"
    _write_rows(
        instruction,
        [
            ["BuildData AI - akceptacja produktów", ""],
            ["Plik źródłowy", source_file],
            ["Liczba produktów", len(products)],
            ["", ""],
            ["Jak używać", "Arkusz 'Produkty' pokazuje produkty jeden po drugim. Uzupełnij status akceptacji i uwagi klienta."],
            ["Poprawki", "W arkuszu 'Cechy produktów' można wpisać poprawioną wartość przy konkretnej cesze produktu."],
            ["Import zwrotny", "Nie zmieniaj kolumn technicznych product_id, attribute_id, parent_attribute_id i row_i."],
        ],
    )
    instruction["A1"].font = Font(bold=True, size=14)

    _product_acceptance_overview_sheet(workbook, products, attribute_labels)
    _product_acceptance_details_sheet(workbook, products, attribute_labels)
    _product_acceptance_json_sheet(workbook, products_payload)

    for sheet in workbook.worksheets:
        _format_sheet(sheet)
        for cell in sheet[1]:
            if cell.value in {"status_akceptacji", "uwagi_klienta", "poprawiona_wartość"}:
                cell.fill = ACCEPT_FILL

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _product_acceptance_overview_sheet(workbook: Workbook, products: list[dict[str, Any]], attribute_labels: dict[int, str]) -> None:
    rows = [[
        "lp",
        "product_id",
        "nazwa_produktu",
        "kod_produktu",
        "liczba_cech",
        "status_akceptacji",
        "uwagi_klienta",
    ]]
    for index, product in enumerate(products, start=1):
        attrs = _product_attrs(product)
        rows.append([
            index,
            product.get("Id", ""),
            _product_identity_value(attrs, attribute_labels, "name"),
            _product_identity_value(attrs, attribute_labels, "code"),
            len([attr for attr in attrs if _attribute_display_value(attr) not in ("", None)]),
            "",
            "",
        ])
    _write_rows(workbook.create_sheet("Produkty"), rows)


def _product_acceptance_details_sheet(workbook: Workbook, products: list[dict[str, Any]], attribute_labels: dict[int, str]) -> None:
    rows = [[
        "lp",
        "product_id",
        "nazwa_produktu",
        "attribute_id",
        "parent_attribute_id",
        "row_i",
        "cecha",
        "wartość",
        "typ_wartości",
        "status_akceptacji",
        "poprawiona_wartość",
        "uwagi_klienta",
    ]]
    for product_index, product in enumerate(products, start=1):
        attrs = _product_attrs(product)
        product_name = _product_identity_value(attrs, attribute_labels, "name")
        for attr in attrs:
            value = _attribute_display_value(attr)
            if value in ("", None):
                continue
            attribute_id = _int_or_text(attr.get("AttributeId"))
            rows.append([
                product_index,
                product.get("Id", ""),
                product_name,
                attribute_id,
                attr.get("ParentAttributeId") or 0,
                attr.get("RowI") or 0,
                attribute_labels.get(attribute_id, f"Atrybut {attribute_id}"),
                value,
                _attribute_value_type(attr),
                "",
                "",
                "",
            ])
    _write_rows(workbook.create_sheet("Cechy produktów"), rows)


def _product_acceptance_json_sheet(workbook: Workbook, products_payload: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("__products_json")
    sheet.sheet_state = "hidden"
    _write_rows(
        sheet,
        [
            ["schema", "builddata.products_acceptance.xlsx.v1"],
            ["products_json", json.dumps(products_payload, ensure_ascii=False, indent=2)],
        ],
    )


def _product_attrs(product: dict[str, Any]) -> list[dict[str, Any]]:
    if isinstance(product.get("productAttributes"), list):
        return product.get("productAttributes") or []
    versions = product.get("dataVersions") or product.get("DataVersions") or []
    if versions and isinstance(versions[0], dict):
        return versions[0].get("productAttributes") or versions[0].get("ProductAttributes") or []
    return []


def _product_identity_value(attrs: list[dict[str, Any]], attribute_labels: dict[int, str], kind: str) -> str:
    preferred_ids = {225, 116, 501} if kind == "name" else {226, 318}
    preferred_words = {"nazwa", "name"} if kind == "name" else {"kod", "code", "pim id", "external"}
    for attr in attrs:
        attribute_id = _int_or_text(attr.get("AttributeId"))
        if attribute_id in preferred_ids:
            value = _attribute_display_value(attr)
            if value:
                return str(value)
    for attr in attrs:
        attribute_id = _int_or_text(attr.get("AttributeId"))
        label = str(attribute_labels.get(attribute_id, "")).lower()
        if any(word in label for word in preferred_words):
            value = _attribute_display_value(attr)
            if value:
                return str(value)
    if kind == "name":
        for attr in attrs:
            if (attr.get("ParentAttributeId") or 0) == 0:
                value = _attribute_display_value(attr)
                if value:
                    return str(value)
    return ""


def _attribute_display_value(attr: dict[str, Any]) -> Any:
    for key in ("varcharValue", "TextValue", "NumberValue", "IntValue", "IntValue2"):
        value = attr.get(key)
        if value not in (None, ""):
            return value
    if attr.get("BooleanValue") is not None:
        return "tak" if bool(attr.get("BooleanValue")) else "nie"
    return ""


def _attribute_value_type(attr: dict[str, Any]) -> str:
    for key, label in (
        ("varcharValue", "tekst"),
        ("TextValue", "długi tekst"),
        ("NumberValue", "liczba"),
        ("IntValue", "liczba całkowita"),
        ("IntValue2", "liczba całkowita 2"),
        ("BooleanValue", "tak/nie"),
    ):
        if attr.get(key) not in (None, ""):
            return label
    return ""


def _int_or_text(value: Any) -> Any:
    try:
        return int(value)
    except (TypeError, ValueError):
        return value


def _mapping_sheet(workbook: Workbook, report: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Mapowanie")
    profile = report.get("product_mapping_profile") or {}
    rows = [[
        "profile_key",
        "source_column",
        "target_path",
        "target_label",
        "target_group",
        "target_value_kind",
        "target_unit",
        "trim",
        "parse_number",
        "decimal_comma",
        "remove_text",
        "replace_from",
        "replace_to",
        "split_by",
        "split_part",
        "unit_source_column",
        "unit_conversion_factor",
        "choice_map",
        "uwagi_klienta",
    ]]
    for profile_key, item in profile.items():
        if str(profile_key).startswith("_") or not isinstance(item, dict):
            continue
        cleanup = item.get("cleanup") or {}
        rows.append([
            profile_key,
            item.get("source_column", ""),
            item.get("target_path", ""),
            item.get("target_label", ""),
            item.get("target_group", ""),
            item.get("target_value_kind", ""),
            item.get("target_unit", ""),
            _yes_no(cleanup.get("trim", True)),
            _yes_no(cleanup.get("parseNumber", False)),
            _yes_no(cleanup.get("decimalComma", False)),
            cleanup.get("removeText", ""),
            cleanup.get("replaceFrom", ""),
            cleanup.get("replaceTo", ""),
            cleanup.get("splitBy", ""),
            cleanup.get("splitPart", ""),
            cleanup.get("unitSourceColumn", ""),
            cleanup.get("unitConversionFactor", ""),
            _format_choice_map(item.get("choice_map") or cleanup.get("choiceMap") or {}),
            "",
        ])
    if len(rows) == 1:
        fallback = report.get("column_mapping") or report.get("system_column_mapping") or {}
        for source_column, target_path in fallback.items():
            rows.append([source_column, source_column, target_path, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""])
    _write_rows(sheet, rows)


def _choice_sheet(workbook: Workbook, report: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Mapy opcji")
    rows = [["profile_key", "source_column", "target_path", "wartość klienta", "opcja PIM", "uwagi_klienta"]]
    for profile_key, item in (report.get("product_mapping_profile") or {}).items():
        if str(profile_key).startswith("_") or not isinstance(item, dict):
            continue
        choice_map = item.get("choice_map") or (item.get("cleanup") or {}).get("choiceMap") or {}
        for source_value, target_value in choice_map.items():
            rows.append([
                profile_key,
                item.get("source_column", ""),
                item.get("target_path", ""),
                source_value,
                target_value,
                "",
            ])
    _write_rows(sheet, rows)


def _row_rules_sheet(workbook: Workbook, report: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Reguły wierszy")
    rows = [["rule_index", "field", "value"]]
    rules = ((report.get("product_mapping_profile") or {}).get("_row_rules") or {}).get("rules") or []
    for index, rule in enumerate(rules, start=1):
        for key, value in rule.items():
            rows.append([index, key, _json_or_text(value)])
    _write_rows(sheet, rows)


def _tables_sheet(workbook: Workbook, report: dict[str, Any]) -> None:
    rows = [["nazwa tabeli", "liczba wierszy", "score produktów", "score elementów"]]
    for table in report.get("tables_detected") or []:
        rows.append([table.get("name", ""), table.get("rows", ""), table.get("product_score", ""), table.get("system_score", "")])
    _write_rows(workbook.create_sheet("Tabele"), rows)


def _unmapped_sheet(workbook: Workbook, report: dict[str, Any]) -> None:
    rows = [["typ", "kolumna"]]
    for column in report.get("unmapped_columns") or []:
        rows.append(["produkty", column])
    for column in report.get("system_unmapped_columns") or []:
        rows.append(["elementy budowlane", column])
    _write_rows(workbook.create_sheet("Nieprzypisane kolumny"), rows)


def _warnings_sheet(workbook: Workbook, report: dict[str, Any]) -> None:
    rows = [["obszar", "wartość"]]
    for key, value in (report.get("warnings") or {}).items():
        if isinstance(value, list):
            if not value:
                rows.append([key, ""])
            for item in value:
                rows.append([key, _json_or_text(item)])
        elif isinstance(value, dict):
            rows.append([key, json.dumps(value, ensure_ascii=False)])
        else:
            rows.append([key, value])
    _write_rows(workbook.create_sheet("Ostrzeżenia"), rows)


def _json_sheet(workbook: Workbook, report: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("__mapping_json")
    sheet.sheet_state = "hidden"
    _write_rows(
        sheet,
        [
            ["schema", "builddata.mapping_report.xlsx.v1"],
            ["product_mapping_profile", json.dumps(report.get("product_mapping_profile") or {}, ensure_ascii=False, indent=2)],
            ["product_mapping", json.dumps(report.get("product_mapping") or {}, ensure_ascii=False, indent=2)],
            ["full_report", json.dumps(report, ensure_ascii=False, indent=2)],
        ],
    )


def _write_rows(sheet: Any, rows: list[list[Any]]) -> None:
    for row in rows:
        sheet.append(row)


def _format_sheet(sheet: Any) -> None:
    if sheet.max_row:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = TECH_FILL if str(cell.value or "").endswith("_key") or cell.value == "target_path" else HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 60)


def _yes_no(value: Any) -> str:
    return "tak" if bool(value) else "nie"


def _format_choice_map(value: Any) -> str:
    if not isinstance(value, dict):
        return ""
    return "\n".join(f"{source} = {target}" for source, target in value.items())


def _json_or_text(value: Any) -> str:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)
