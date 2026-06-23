from __future__ import annotations

import csv
import hashlib
import json
import re
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from urllib.parse import unquote, urlparse
from typing import Any

from openpyxl import load_workbook

from .mapping import (
    analyze_tables,
    apply_column_mapping,
    apply_mapping_profile_to_rows,
    attribute_order,
    field_definitions_payload,
    int_value,
    is_deleted,
    is_nested_attribute,
    is_type_series_parent,
    load_json_content,
    pim_bundle_file_key,
    pim_items,
    product_fields_from_json,
    product_fields_from_pim_bundle,
    product_model_choices_from_pim_bundle,
    semantic_pim_field_key,
    skip_pim_attribute,
    value_kind_from_attribute,
)
from .report_export import mapping_report_xlsx_bytes, product_acceptance_xlsx_bytes


PRODUCT_MODEL_TYPE = 66
PRODUCT_TYPE_ID = 1
PRODUCT_ID_START = 900000
BUILDING_ELEMENT_MODEL_TYPE = 74
BUILDING_ELEMENT_TYPE_ID = 1
BUILDING_ELEMENT_ID_START = 910000
COLOR_TYPE_ID = 1
COLOR_ID_START = 920000


PIM_ATTR = {
    "product_name": 225,
    "external_id": 226,
    "old_product_id": 227,
    "available": 228,
    "new": 229,
    "categories": 230,
    "unit": 231,
    "product_information": 233,
    "packages": 234,
    "palettes": 235,
    "variants": 236,
    "documents": 237,
    "sot_table": 276,
}

PRODUCT_INFO_ATTR = {
    "short_name": 243,
    "full_name": 244,
    "technical_name": 245,
    "description": 246,
    "properties": 247,
    "prosperation": 248,
    "surface_preparation": 249,
    "usage_method": 250,
    "comments": 251,
    "warnings": 252,
    "norms": 253,
    "storage": 254,
    "manufacturer": 302,
    "url": 303,
    "package_text": 304,
}

PRODUCT_INFO_FIELD_TARGETS = {
    "description": "product.description.value",
    "properties": "product.properties.value",
    "prosperation": "product.application.value",
    "usage_method": "product.usage_method.value",
    "norms": "product.norms.value",
    "manufacturer": "product.manufacturer.value",
    "url": "product.product_url.value",
}

PACKAGE_ATTR = {
    "name": 255,
    "weight": 256,
    "capacity": 257,
    "unit_name": 258,
}

SOT_ATTR = {
    "thickness": 277,
    "lambda": 278,
    "density": 279,
    "vapor_permeability": 295,
}

SOT_FIELD_TARGETS = {
    "thickness": "type_series[].thickness.value",
    "lambda": "type_series[].lambda_value.value",
    "density": "type_series[].density.value",
    "vapor_permeability": "type_series[].vapor_permeability_mu.value",
}

DOCUMENT_ATTR = {
    "name": 270,
    "display_web": 271,
    "category": 272,
    "url": 273,
    "extension": 274,
    "safe_name": 275,
}

GENERIC_PIM_PARENT_IDS = {
    **{attribute_id: PIM_ATTR["product_information"] for attribute_id in [243, 244, 245, 249, 251, 252, 254, 302, 303, 304]},
    **{attribute_id: PIM_ATTR["packages"] for attribute_id in [255, 256, 257, 258]},
    **{attribute_id: PIM_ATTR["palettes"] for attribute_id in [259, 260, 261, 262]},
    **{attribute_id: PIM_ATTR["variants"] for attribute_id in [263, 264, 265, 266, 267, 268, 269]},
    **{attribute_id: PIM_ATTR["documents"] for attribute_id in [270, 271, 272, 273, 274, 275]},
}

BUILDING_ELEMENT_ATTR = {
    "name": 280,
    "building_element_type": 281,
    "variants": 283,
    "variant_name": 284,
    "layers": 285,
    "layer_position": 286,
    "layer_name": 287,
    "available_products": 289,
    "product": 290,
    "insulation_type": 292,
    "default": 298,
    "bim_building_element_type": 299,
}

DEFAULT_COLOR_PARAMETERS = [
    {"Name": "name", "DispName": "Name", "type": "VarChar", "options": None, "unit": None},
    {"Name": "type", "DispName": "Type", "type": "Select", "options": [{"name": "Color", "value": "simple"}, {"name": "Texture", "value": "advanced"}], "unit": None},
    {"Name": "description", "DispName": "Description", "type": "Longtext", "options": None, "unit": None},
    {"Name": "r", "DispName": "R", "type": "Int", "options": None, "unit": None},
    {"Name": "g", "DispName": "G", "type": "Int", "options": None, "unit": None},
    {"Name": "b", "DispName": "B", "type": "Int", "options": None, "unit": None},
    {"Name": "colorRGB", "DispName": "RGB / HEX", "type": "colorRGB", "options": None, "unit": None},
    {"Name": "hbw", "DispName": "HBW", "type": "Int", "options": None, "unit": None},
    {"Name": "RealWidth", "DispName": "Real width of loaded texture", "type": "Number", "options": None, "unit": "m"},
    {"Name": "MainTexture", "DispName": "Main Texture", "type": "Files", "options": None, "unit": None},
    {"Name": "Thumbnail", "DispName": "Thumbnail of the texture", "type": "Files", "options": None, "unit": None},
    {"Name": "base_color_map", "DispName": "Base Color Map", "type": "Files", "options": None, "unit": None},
    {"Name": "normal_map", "DispName": "Normal Map", "type": "Files", "options": None, "unit": None},
    {"Name": "displacement_map", "DispName": "Displacement Map", "type": "Files", "options": None, "unit": None},
    {"Name": "opacity_map", "DispName": "Opacity Map", "type": "Files", "options": None, "unit": None},
    {"Name": "roughness", "DispName": "Roughness", "type": "Select", "options": [{"name": "Glossy", "value": "glossy"}, {"name": "Matte", "value": "matte"}, {"name": "Roughness Map", "value": "roughnessMap"}], "unit": None},
    {"Name": "roughness_map", "DispName": "Roughness Map", "type": "Files", "options": None, "unit": None},
]

FIELD_TYPE_STRING_OPTION_ID = 265


@dataclass(frozen=True)
class PimExportSchema:
    product_model_id: int = PRODUCT_MODEL_TYPE
    strict_model: bool = False
    product_attribute_ids: dict[str, int] | None = None
    product_parent_by_attribute_id: dict[int, int] | None = None
    attribute_value_kinds: dict[int, str] | None = None
    known_attribute_ids: frozenset[int] = frozenset()
    type_series_parent_id: int | None = None
    type_series_attribute_ids: dict[str, int] | None = None
    type_series_parent_by_attribute_id: dict[int, int] | None = None

    def product_attribute_id(self, target_path: str, fallback: int | None = None) -> int | None:
        attribute_id = (self.product_attribute_ids or {}).get(target_path)
        if attribute_id is not None:
            return attribute_id
        if fallback is not None and (not self.strict_model or fallback in self.known_attribute_ids):
            return fallback
        return None

    def static_attribute_id(self, fallback: int) -> int | None:
        if not self.strict_model or fallback in self.known_attribute_ids:
            return fallback
        return None

    def has_attribute(self, attribute_id: int) -> bool:
        return not self.strict_model or attribute_id in self.known_attribute_ids

    def attribute_value_kind(self, attribute_id: int, fallback: str = "free_text") -> str:
        return (self.attribute_value_kinds or {}).get(attribute_id, fallback)

    def product_parent_for_attribute(self, attribute_id: int | None = None, fallback: int = 0) -> int:
        if attribute_id is not None and attribute_id in (self.product_parent_by_attribute_id or {}):
            return (self.product_parent_by_attribute_id or {})[attribute_id]
        if self.strict_model:
            return 0
        return fallback

    def type_series_attribute_id(self, target_path: str) -> int | None:
        return (self.type_series_attribute_ids or {}).get(target_path)

    def type_series_parent_for_attribute(self, attribute_id: int | None = None) -> int:
        if attribute_id is not None:
            parent_id = (self.type_series_parent_by_attribute_id or {}).get(attribute_id)
            if parent_id:
                return parent_id
        return self.type_series_parent_id or PIM_ATTR["sot_table"]


DEFAULT_EXPORT_SCHEMA = PimExportSchema(
    product_model_id=PRODUCT_MODEL_TYPE,
    product_attribute_ids={
        "product.name.value": PIM_ATTR["product_name"],
        "product.code.value": PIM_ATTR["external_id"],
        "product.unit.value": PIM_ATTR["unit"],
        "product.category[].value": PIM_ATTR["categories"],
        "product.description.value": PRODUCT_INFO_ATTR["description"],
        "product.properties.value": PRODUCT_INFO_ATTR["properties"],
        "product.application.value": PRODUCT_INFO_ATTR["prosperation"],
        "product.usage_method.value": PRODUCT_INFO_ATTR["usage_method"],
        "product.norms.value": PRODUCT_INFO_ATTR["norms"],
        "product.manufacturer.value": PRODUCT_INFO_ATTR["manufacturer"],
        "product.product_url.value": PRODUCT_INFO_ATTR["url"],
    },
    product_parent_by_attribute_id={
        PRODUCT_INFO_ATTR["description"]: PIM_ATTR["product_information"],
        PRODUCT_INFO_ATTR["properties"]: PIM_ATTR["product_information"],
        PRODUCT_INFO_ATTR["prosperation"]: PIM_ATTR["product_information"],
        PRODUCT_INFO_ATTR["usage_method"]: PIM_ATTR["product_information"],
        PRODUCT_INFO_ATTR["norms"]: PIM_ATTR["product_information"],
        PRODUCT_INFO_ATTR["manufacturer"]: PIM_ATTR["product_information"],
        PRODUCT_INFO_ATTR["url"]: PIM_ATTR["product_information"],
    },
    attribute_value_kinds={
        PIM_ATTR["product_name"]: "free_text",
        PIM_ATTR["external_id"]: "free_text",
        PIM_ATTR["unit"]: "free_text",
        PIM_ATTR["categories"]: "multi_choice",
        PRODUCT_INFO_ATTR["description"]: "free_text",
        PRODUCT_INFO_ATTR["properties"]: "free_text",
        PRODUCT_INFO_ATTR["prosperation"]: "free_text",
        PRODUCT_INFO_ATTR["usage_method"]: "free_text",
        PRODUCT_INFO_ATTR["norms"]: "free_text",
        PRODUCT_INFO_ATTR["manufacturer"]: "free_text",
        PRODUCT_INFO_ATTR["url"]: "free_text",
        SOT_ATTR["thickness"]: "number",
        SOT_ATTR["lambda"]: "number",
        SOT_ATTR["density"]: "number",
        SOT_ATTR["vapor_permeability"]: "number",
    },
    known_attribute_ids=frozenset(),
    type_series_parent_id=PIM_ATTR["sot_table"],
    type_series_attribute_ids={
        "type_series[].thickness.value": SOT_ATTR["thickness"],
        "type_series[].lambda_value.value": SOT_ATTR["lambda"],
        "type_series[].density.value": SOT_ATTR["density"],
        "type_series[].vapor_permeability_mu.value": SOT_ATTR["vapor_permeability"],
    },
    type_series_parent_by_attribute_id={
        SOT_ATTR["thickness"]: PIM_ATTR["sot_table"],
        SOT_ATTR["lambda"]: PIM_ATTR["sot_table"],
        SOT_ATTR["density"]: PIM_ATTR["sot_table"],
        SOT_ATTR["vapor_permeability"]: PIM_ATTR["sot_table"],
    },
)


CATEGORY_OPTIONS = {
    "fast inneo": 243,
    "farby": 244,
    "farby elewacyjne": 245,
    "farby i wyprawy gruntujace": 246,
    "farby i wyprawy gruntujące": 246,
    "farby specjalne": 247,
    "farby wewnetrzne": 248,
    "farby wewnętrzne": 248,
    "kleje do plytek ceramicznych": 249,
    "kleje do płytek ceramicznych": 249,
    "kleje do styropianu": 250,
    "kleje do welny mineralnej": 251,
    "kleje do wełny mineralnej": 251,
    "kleje dyspersyjne": 252,
    "masy szpachlowe i wyrownujace": 253,
    "masy szpachlowe i wyrównujące": 253,
    "posadzki samopoziomujace i podkladowe": 254,
    "posadzki samopoziomujące i podkładowe": 254,
    "preparaty gruntujace i ochronne": 255,
    "preparaty gruntujące i ochronne": 255,
    "srodki hydroizolacyjne": 256,
    "środki hydroizolacyjne": 256,
    "tynki mozaikowe": 257,
    "tynki polimerowo-mineralne": 258,
    "tynki silikatowe": 259,
    "tynki silikonowe": 260,
    "tynki siloksanowe": 261,
    "zaprawy klejowe": 262,
    "zaprawy murarskie i tynkarskie": 263,
}

BUILDING_ELEMENT_TYPE_OPTIONS = {
    "ocieplenie scian od zewnatrz": 276,
    "ocieplenie ścian od zewnątrz": 276,
    "ocieplenie scian od wewnatrz": 277,
    "ocieplenie ścian od wewnątrz": 277,
    "dodatkowe ocieplenie na istniejacym ociepleniu": 278,
    "dodatkowe ocieplenie na istniejącym ociepleniu": 278,
    "ocieplenie stropu garazowego": 309,
    "ocieplenie stropu garażowego": 309,
}

INSULATION_TYPE_OPTIONS = {
    "welna mineralna": 292,
    "wełna mineralna": 292,
    "styropian": 293,
}

BIM_TYPE_OPTIONS = {
    "wall": 310,
    "sciana": 310,
    "ściana": 310,
    "floor": 311,
    "podloga": 311,
    "podłoga": 311,
    "ceiling": 312,
    "strop": 312,
    "roof": 313,
    "dach": 313,
    "curtain wall": 314,
    "other": 315,
    "inne": 315,
}

LAYER_POSITION_OPTIONS = {
    1: 282,
    2: 283,
    3: 284,
    4: 285,
    5: 286,
    6: 287,
    7: 288,
    8: 289,
    9: 290,
    10: 291,
    11: 303,
}


CANONICAL_ALIASES = {
    "product_name": {
        "product_name",
        "name",
        "nazwa",
        "nazwa produktu",
        "produkt",
        "product",
        "model",
    },
    "external_id": {
        "id",
        "sku",
        "kod",
        "kod produktu",
        "indeks",
        "external_id",
        "product id",
    },
    "unit": {"unit", "jednostka", "jm", "jednostka miary"},
    "categories": {"category", "categories", "kategoria", "kategorie", "grupa", "parent_group", "parent path", "parent_path", "group id", "category id"},
    "short_name": {"short_name", "skrot", "skrót", "model handlowy"},
    "full_name": {"full_name", "pelna nazwa", "pełna nazwa", "nazwa pelna"},
    "technical_name": {"technical_name", "nazwa techniczna", "technical name"},
    "description": {"description", "opis", "opis produktu"},
    "properties": {"properties", "wlasciwosci", "właściwości", "cechy"},
    "prosperation": {"zastosowanie", "application", "use", "przeznaczenie"},
    "usage_method": {"sposob uzycia", "sposób użycia", "usage_method", "aplikacja"},
    "norms": {"normy", "norms", "certyfikaty", "aprobaty"},
    "manufacturer": {"manufacturer", "producent"},
    "url": {"url", "link", "www", "strona"},
    "package_text": {"opakowanie", "package", "packaging"},
    "documents": {"files", "file", "file links", "document urls", "document_urls", "image urls", "image_urls", "packshot", "packshot_url", "url dokumentu", "link dokumentu"},
    "weight": {"waga", "weight", "masa"},
    "capacity": {"pojemnosc", "pojemność", "capacity"},
    "thickness": {"grubosc", "grubość", "thickness", "sot_thickness"},
    "lambda": {"lambda", "λ", "wspolczynnik lambda", "współczynnik lambda"},
    "density": {"gestosc", "gęstość", "density"},
    "vapor_permeability": {"mu", "μ", "vapor permeability", "paroprzepuszczalnosc"},
}

SYSTEM_ALIASES = {
    "system_name": {"system_name", "system", "nazwa systemu", "nazwa elementu", "building element", "element"},
    "building_element_type": {"building_element_type", "typ systemu", "zastosowanie systemu", "typ elementu"},
    "insulation_type": {"insulation_type", "rodzaj izolacji", "izolacja"},
    "bim_building_element_type": {"bim_building_element_type", "bim type", "typ bim", "building element type bim"},
    "variant_name": {"variant_name", "wariant", "nazwa wariantu", "finish", "wykonczenie", "wykończenie"},
    "layer_position": {"layer_position", "pozycja warstwy", "kolejnosc", "kolejność", "nr warstwy"},
    "layer_name": {"layer_name", "warstwa", "nazwa warstwy"},
    "product_code": {"product_code", "kod produktu", "sku", "indeks", "id produktu"},
    "product_name": {"product_name", "produkt", "nazwa produktu"},
    "default": {"default", "domyslny", "domyślny", "produkt domyslny", "produkt domyślny"},
    "quantity": {"quantity", "ilosc", "ilość", "zuzycie", "zużycie"},
}


@dataclass(frozen=True)
class SourceTable:
    name: str
    rows: list[dict[str, Any]]


def analyze_uploaded_file(
    filename: str,
    content: bytes,
    *,
    product_model_content: bytes | None = None,
    product_model_files: dict[str, bytes] | None = None,
    product_root_model_id: int | None = None,
) -> dict[str, Any]:
    tables = read_source_tables(filename, content)
    product_fields = None
    product_root_model_id = int_value(product_root_model_id)
    if product_model_files:
        product_fields = product_fields_from_pim_bundle(product_model_files, root_model_id=product_root_model_id)
    elif product_model_content:
        product_fields = product_fields_from_json(product_model_content)
    return analyze_tables(tables, product_fields=product_fields)


def analyze_product_model_files(product_model_files: dict[str, bytes], product_root_model_id: int | None = None) -> dict[str, Any]:
    product_root_model_id = int_value(product_root_model_id)
    model_choices = product_model_choices_from_pim_bundle(product_model_files)
    selected_root_model_id = product_root_model_id or (model_choices[0]["id"] if model_choices else None)
    fields = product_fields_from_pim_bundle(product_model_files, root_model_id=selected_root_model_id)
    if not fields:
        raise ValueError("Product model files do not contain a readable PIM product model.")
    return {
        "status": "model_loaded",
        "files": sorted(product_model_files.keys()),
        "selected_root_model_id": selected_root_model_id,
        "product_models": model_choices,
        "target_fields": field_definitions_payload(fields),
    }


def color_parameters_from_content(content: bytes | str | None = None) -> list[dict[str, Any]]:
    if not content:
        return [dict(item) for item in DEFAULT_COLOR_PARAMETERS]
    payload = load_json_content(content)
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict) and item.get("Name")]
    if isinstance(payload, dict):
        for key in ("parameters", "colorParameters", "items"):
            value = payload.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict) and item.get("Name")]
    return [dict(item) for item in DEFAULT_COLOR_PARAMETERS]


def color_fields_payload(parameters: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for parameter in parameters:
        name = str(parameter.get("Name") or "").strip()
        if not name:
            continue
        param_type = str(parameter.get("type") or parameter.get("Type") or "VarChar").strip()
        result.append(
            {
                "key": name,
                "label": str(parameter.get("DispName") or name).strip() or name,
                "type": param_type,
                "unit": parameter.get("unit"),
                "options": parameter.get("options") or [],
                "is_file": normalize_lookup(param_type) == "files",
            }
        )
    return result


def analyze_colors_file(filename: str, content: bytes, color_parameters_content: bytes | str | None = None) -> dict[str, Any]:
    tables = read_source_tables(filename, content)
    parameters = color_parameters_from_content(color_parameters_content)
    return {
        "status": "colors_analyzed",
        "filename": filename,
        "fields": color_fields_payload(parameters),
        "tables": [
            {
                "name": table.name,
                "rows": len(table.rows),
                "columns": columns_for_rows(table.rows),
                "column_values": distinct_column_values(table.rows),
                "sample_rows": table.rows[:5],
            }
            for table in tables
        ],
    }


def convert_colors_file(
    filename: str,
    content: bytes,
    output_root: Path,
    *,
    color_mapping: dict[str, str] | None = None,
    color_choice_mapping: dict[str, dict[str, str]] | None = None,
    color_parameters_content: bytes | str | None = None,
    table_name: str | None = None,
) -> dict[str, Any]:
    tables = read_source_tables(filename, content)
    if not tables:
        raise ValueError("Uploaded file does not contain readable color rows.")
    table = next((item for item in tables if item.name == table_name), tables[0])
    parameters = color_parameters_from_content(color_parameters_content)
    parameter_by_name = {str(item.get("Name") or ""): item for item in parameters if item.get("Name")}
    mapping = color_mapping or {}
    colors = [
        build_color_entry(row, index, mapping, parameter_by_name, color_choice_mapping or {})
        for index, row in enumerate(table.rows, start=1)
        if any(value not in (None, "") for value in row.values())
    ]
    payload = {"Count": len(colors), "colors": colors}
    job_id = make_job_id(f"colors:{filename}", content)
    output_dir = output_root / job_id
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "colors.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    report = {
        "source_filename": filename,
        "table": table.name,
        "source_rows": len(table.rows),
        "colors_generated": len(colors),
        "file_parameters_are_references": True,
        "mapped_fields": {key: value for key, value in mapping.items() if value},
        "choice_mapping": color_choice_mapping or {},
    }
    (output_dir / "colors_mapping_report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "status": "colors_converted",
        "job_id": job_id,
        "colors_count": len(colors),
        "files": {
            "colors_json": f"/outputs/{job_id}/colors.json",
            "mapping_report_json": f"/outputs/{job_id}/colors_mapping_report.json",
        },
        "report": report,
    }


def columns_for_rows(rows: list[dict[str, Any]]) -> list[str]:
    columns: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for column in row:
            if column not in seen:
                seen.add(column)
                columns.append(column)
    return columns


def distinct_column_values(rows: list[dict[str, Any]], limit: int = 60) -> dict[str, list[str]]:
    values: dict[str, list[str]] = {}
    seen: dict[str, set[str]] = {}
    for row in rows:
        for column, value in row.items():
            if value in (None, ""):
                continue
            text = str(value).strip()
            if not text:
                continue
            column_seen = seen.setdefault(column, set())
            if text in column_seen or len(column_seen) >= limit:
                continue
            column_seen.add(text)
            values.setdefault(column, []).append(text)
    return values


def build_color_entry(
    row: dict[str, Any],
    index: int,
    mapping: dict[str, str],
    parameter_by_name: dict[str, dict[str, Any]],
    choice_mapping: dict[str, dict[str, str]],
) -> dict[str, Any]:
    color_id = COLOR_ID_START + index
    parameters: list[dict[str, Any]] = []
    files_parameters: list[dict[str, Any]] = []
    rgb_value = mapped_row_value(row, mapping.get("colorRGB"))
    rgb = parse_rgb_value(rgb_value)
    for component, value in zip(("r", "g", "b"), rgb or (None, None, None)):
        if value is not None and not mapped_row_value(row, mapping.get(component)):
            add_color_parameter(parameters, color_id, component, value, parameter_by_name.get(component, {"type": "Int"}))
    for parameter_name, source_column in mapping.items():
        if not source_column or parameter_name == "colorRGB":
            continue
        value = mapped_row_value(row, source_column)
        if value in (None, ""):
            continue
        parameter_def = parameter_by_name.get(parameter_name, {"Name": parameter_name, "type": "VarChar"})
        if normalize_lookup(parameter_def.get("type")) == "files":
            files_parameters.extend(color_file_parameters(color_id, parameter_name, value))
            continue
        add_color_parameter(parameters, color_id, parameter_name, value, parameter_def, choice_mapping.get(parameter_name, {}))
    if not any(item["parameterName"] == "type" for item in parameters):
        file_mode = bool(files_parameters)
        add_color_parameter(parameters, color_id, "type", "advanced" if file_mode else "simple", parameter_by_name.get("type", {"type": "Select"}), choice_mapping.get("type", {}))
    return {
        "Id": color_id,
        "TypeId": COLOR_TYPE_ID,
        "dataVersions": [
            {
                "VersionId": 1,
                "parameters": parameters,
                "filesParameters": files_parameters,
            }
        ],
    }


def mapped_row_value(row: dict[str, Any], source_column: str | None) -> Any:
    if not source_column:
        return None
    return row.get(source_column)


def add_color_parameter(
    parameters: list[dict[str, Any]],
    color_id: int,
    name: str,
    value: Any,
    definition: dict[str, Any],
    choice_map: dict[str, str] | None = None,
) -> None:
    param_type = normalize_lookup(definition.get("type") or definition.get("Type") or "VarChar")
    item = {
        "ColorId": color_id,
        "VersionId": 1,
        "parameterName": name,
        "TextValue": None,
        "varcharValue": None,
        "IntValue": None,
        "NumberValue": None,
        "BooleanValue": False,
    }
    if param_type in {"int", "integer"}:
        item["IntValue"] = int_value(value)
    elif param_type in {"number", "float", "decimal"}:
        item["NumberValue"] = parse_number(value)
    elif param_type in {"select", "checkboxes"}:
        item["TextValue"] = normalize_color_option(value, definition, choice_map or {})
    elif param_type in {"longtext", "text", "textarea"}:
        item["TextValue"] = str(value).strip()
    else:
        item["varcharValue"] = str(value).strip()
    if item["TextValue"] is not None or item["varcharValue"] is not None or item["IntValue"] is not None or item["NumberValue"] is not None:
        parameters.append(item)


def normalize_color_option(value: Any, definition: dict[str, Any], choice_map: dict[str, str] | None = None) -> str:
    raw = str(value or "").strip()
    if choice_map:
        mapped = choice_map.get(raw) or choice_map.get(normalize_lookup(raw))
        if mapped:
            return str(mapped).strip()
    raw_key = normalize_lookup(raw)
    for option in definition.get("options") or []:
        if not isinstance(option, dict):
            continue
        if raw_key in {normalize_lookup(option.get("value")), normalize_lookup(option.get("name"))}:
            return str(option.get("value") or option.get("name") or raw).strip()
    if raw_key in {"color", "kolor", "simple", "prosty"}:
        return "simple"
    if raw_key in {"texture", "tekstura", "advanced", "zlozony", "złożony"}:
        return "advanced"
    return raw


def color_file_parameters(color_id: int, parameter_name: str, value: Any) -> list[dict[str, Any]]:
    values = split_multi_value(value)
    result = []
    for order, file_ref in enumerate(values):
        text = str(file_ref).strip()
        if not text:
            continue
        is_url = text.lower().startswith(("http://", "https://"))
        result.append(
            {
                "Id": color_id * 1000 + len(result) + 1,
                "ElementId": color_id,
                "VersionId": 1,
                "parameterName": parameter_name,
                "dispName": None,
                "fileName": None if is_url else Path(text).name,
                "fileUrl": text if is_url else None,
                "OrderDisplayOrder": order,
            }
        )
    return result


def split_multi_value(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    text = str(value or "").strip()
    if not text:
        return []
    return [item.strip() for item in re.split(r"[;\n|]+", text) if item.strip()]


def parse_rgb_value(value: Any) -> tuple[int, int, int] | None:
    text = str(value or "").strip()
    if not text:
        return None
    hex_match = re.fullmatch(r"#?([0-9a-fA-F]{6})", text)
    if hex_match:
        raw = hex_match.group(1)
        return int(raw[0:2], 16), int(raw[2:4], 16), int(raw[4:6], 16)
    numbers = [int_value(item) for item in re.findall(r"\d+", text)[:3]]
    if len(numbers) == 3 and all(item is not None for item in numbers):
        return tuple(max(0, min(255, int(item))) for item in numbers)  # type: ignore[return-value]
    return None


def convert_uploaded_file(
    filename: str,
    content: bytes,
    output_root: Path,
    *,
    product_mapping: dict[str, str] | None = None,
    product_mapping_profile: dict[str, Any] | None = None,
    system_mapping: dict[str, str] | None = None,
) -> dict[str, Any]:
    tables = read_source_tables(filename, content)
    product_rows = choose_product_rows(tables)
    system_rows = choose_system_rows(tables)
    if product_mapping_profile:
        product_rows = apply_mapping_profile_to_rows(product_rows, product_mapping_profile)
    elif product_mapping:
        product_rows = [apply_column_mapping(row, product_mapping) for row in product_rows]
    if system_mapping:
        system_rows = [apply_column_mapping(row, system_mapping) for row in system_rows]
    mapped_rows = unique_mapped_products([map_source_row(row) for row in product_rows])
    products = [build_pim_product(mapped, index) for index, mapped in enumerate(mapped_rows, start=1)]
    product_index = build_product_index(products)
    system_mapped_rows = [map_system_row(row) for row in system_rows]
    building_elements, system_warnings = build_building_elements(system_mapped_rows, product_index)
    report = build_mapping_report(
        filename,
        tables,
        product_rows,
        mapped_rows,
        products,
        system_rows=system_rows,
        system_mapped_rows=system_mapped_rows,
        building_elements=building_elements,
        system_warnings=system_warnings,
    )

    job_id = make_job_id(filename, content)
    output_dir = output_root / job_id
    output_dir.mkdir(parents=True, exist_ok=True)

    products_payload = {
        "productsCount": len(products),
        "products": products,
    }
    (output_dir / "products.json").write_text(
        json.dumps(products_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    building_elements_payload = {
        "buildingElementsCount": len(building_elements),
        "buildingElements": building_elements,
    }
    (output_dir / "building_elements.json").write_text(
        json.dumps(building_elements_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (output_dir / "mapping_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return {
        "job_id": job_id,
        "status": "converted",
        "products_count": len(products),
        "building_elements_count": len(building_elements),
        "output_dir": str(output_dir),
        "files": {
            "products_json": f"/outputs/{job_id}/products.json",
            "building_elements_json": f"/outputs/{job_id}/building_elements.json",
            "mapping_report_json": f"/outputs/{job_id}/mapping_report.json",
        },
        "report": report,
    }


def canonical_product_key(row: dict[str, Any]) -> str:
    for key in ("product.code.value", "product.name.value"):
        value = row.get(key)
        if value not in (None, ""):
            return normalize_lookup(value)
    return ""


def should_apply_enrichment(row: dict[str, Any], entry: dict[str, Any]) -> bool:
    scope = str(entry.get("scope") or "all")
    if scope == "all":
        return True
    if scope == "current_product":
        return canonical_product_key(row) == normalize_lookup(entry.get("product_key"))
    return False


def value_is_empty(value: Any) -> bool:
    if value in (None, ""):
        return True
    if isinstance(value, list) and not value:
        return True
    return False


def type_series_row_index_from_entry(entry: dict[str, Any]) -> int | None:
    raw_value = entry.get("type_series_row_index")
    if raw_value in (None, ""):
        return None
    try:
        row_index = int(raw_value)
    except (TypeError, ValueError):
        return None
    return row_index if row_index >= 0 else None


def apply_enrichment_value(
    row: dict[str, Any],
    target_path: str,
    value: Any,
    mode: str,
    type_series_row_index: int | None = None,
) -> bool:
    if not target_path or target_path == "ignore" or value in (None, ""):
        return False
    replace = mode == "replace"
    if target_path.startswith("type_series[]"):
        variants = row.get("_type_series_rows")
        if not isinstance(variants, list):
            variants = []
            row["_type_series_rows"] = variants
        if not variants:
            variants.append({})
        changed = False
        if type_series_row_index is not None:
            if type_series_row_index >= len(variants):
                return False
            target_variants = [variants[type_series_row_index]]
        else:
            target_variants = variants
        for variant in target_variants:
            if replace or value_is_empty(variant.get(target_path)):
                variant[target_path] = value
                changed = True
        return changed
    if replace or value_is_empty(row.get(target_path)):
        row[target_path] = value
        return True
    return False


def apply_enrichment_session_to_rows(
    rows: list[dict[str, Any]],
    enrichment_session: dict[str, Any] | None,
) -> dict[str, Any]:
    if not enrichment_session:
        return {}
    entries = [
        entry
        for entry in enrichment_session.get("manual_entries") or []
        if isinstance(entry, dict) and entry.get("target_path") and entry.get("value") not in (None, "")
    ]
    applied: list[dict[str, Any]] = []
    for entry in entries:
        target_path = str(entry.get("target_path") or "")
        mode = str(entry.get("mode") or "missing_only")
        type_series_row_index = None
        if target_path.startswith("type_series[]") and entry.get("apply_type_series_to_all") is False:
            type_series_row_index = type_series_row_index_from_entry(entry)
        for index, row in enumerate(rows, start=1):
            if not should_apply_enrichment(row, entry):
                continue
            if apply_enrichment_value(row, target_path, entry.get("value"), mode, type_series_row_index):
                applied.append(
                    {
                        "source": entry.get("source") or "manual",
                        "target_path": target_path,
                        "product_index": index,
                        "product_key": canonical_product_key(row),
                        "mode": mode,
                        "type_series_row_index": type_series_row_index,
                    }
                )
    return {
        "manual_entries_count": len(entries),
        "typical_sources_count": len(enrichment_session.get("typical_sources") or []),
        "applied_count": len(applied),
        "applied": applied,
    }


def products_from_pim_payload(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if isinstance(payload, dict) and isinstance(payload.get("products"), list):
        return [item for item in payload["products"] if isinstance(item, dict)]
    if isinstance(payload, dict) and isinstance(payload.get("Products"), list):
        return [item for item in payload["Products"] if isinstance(item, dict)]
    return []


def product_attributes(product: dict[str, Any]) -> list[dict[str, Any]]:
    attrs = product.get("productAttributes") or product.get("ProductAttributes")
    if isinstance(attrs, list):
        return attrs
    versions = product.get("dataVersions") or product.get("DataVersions")
    if not isinstance(versions, list) or not versions:
        return []
    attrs = versions[0].get("productAttributes") or versions[0].get("ProductAttributes")
    return attrs if isinstance(attrs, list) else []


def attr_value(attr: dict[str, Any]) -> Any:
    for key in ("varcharValue", "TextValue", "NumberValue", "IntValue", "IntValue2", "BooleanValue"):
        value = attr.get(key)
        if value not in (None, ""):
            return value
    return None


def product_identity_keys(product: dict[str, Any]) -> set[str]:
    keys: set[str] = set()
    for attr in product_attributes(product):
        is_known_identity = attr.get("AttributeId") in {PIM_ATTR["product_name"], PIM_ATTR["external_id"]}
        is_top_level_text = (attr.get("ParentAttributeId") or 0) == 0 and any(attr.get(key) for key in ("varcharValue", "TextValue"))
        if not is_known_identity and not is_top_level_text:
            continue
        value = attr_value(attr)
        normalized = normalize_lookup(value)
        if normalized:
            keys.add(normalized)
    return keys


def attr_identity(attr: dict[str, Any]) -> tuple[Any, Any, Any]:
    return (attr.get("AttributeId"), attr.get("ParentAttributeId") or 0, attr.get("RowI") or 0)


def attr_has_import_value(attr: dict[str, Any]) -> bool:
    return attr_value(attr) not in (None, "")


def attr_is_type_series_row(attr: dict[str, Any]) -> bool:
    return (attr.get("ParentAttributeId") or 0) == PIM_ATTR["sot_table"]


def type_series_row_indices(attrs: list[dict[str, Any]]) -> list[int]:
    indices = {
        int(attr.get("RowI") or 0)
        for attr in attrs
        if attr_is_type_series_row(attr) and int(attr.get("RowI") or 0) > 0
    }
    return sorted(indices)


def explicit_typical_match_map(enrichment_session: dict[str, Any] | None) -> dict[str, list[dict[str, Any]]]:
    if not enrichment_session:
        return {}
    matches: dict[str, list[dict[str, Any]]] = {}
    for match in enrichment_session.get("typical_matches") or []:
        if not isinstance(match, dict):
            continue
        product_key = normalize_lookup(match.get("product_key"))
        typical_key = normalize_lookup(match.get("typical_key"))
        if product_key and typical_key:
            matches.setdefault(product_key, []).append(
                {
                    "typical_key": typical_key,
                    "selected_attributes": set(match.get("selected_attributes") or []),
                    "has_selected_attributes": match.get("source") == "typical_manual_match" or "selected_attributes" in match,
                    "apply_type_series_to_all": match.get("apply_type_series_to_all") is not False,
                }
            )
    return matches


def typical_attribute_target(attr: dict[str, Any], enrichment_session: dict[str, Any] | None) -> dict[str, Any] | None:
    if not enrichment_session:
        return dict(attr)
    source_key = f"{attr.get('AttributeId')}:{attr.get('ParentAttributeId') or 0}"
    source_key_with_row = f"{source_key}:{attr.get('RowI') or 0}"
    attribute_map = enrichment_session.get("typical_attribute_map") or {}
    target = attribute_map.get(source_key_with_row) or attribute_map.get(source_key)
    if target is None:
        return dict(attr)
    target_attribute_id = target.get("attribute_id")
    if not target_attribute_id:
        return None
    copied = dict(attr)
    copied["AttributeId"] = target_attribute_id
    copied["ParentAttributeId"] = target.get("parent_attribute_id") or 0
    copied["RowI"] = target.get("row_i") if target.get("row_i") is not None else (attr.get("RowI") or 0)
    copied["hash"] = None
    copied["parentHash"] = ""
    return copied


def apply_typical_products_to_products(
    products: list[dict[str, Any]],
    typical_products_payload: Any,
    enrichment_session: dict[str, Any] | None = None,
) -> dict[str, Any]:
    typical_products = products_from_pim_payload(typical_products_payload)
    if not typical_products:
        return {}
    typical_by_key: dict[str, dict[str, Any]] = {}
    for typical in typical_products:
        for key in product_identity_keys(typical):
            typical_by_key.setdefault(key, typical)
    explicit_matches = explicit_typical_match_map(enrichment_session)

    applied: list[dict[str, Any]] = []
    for index, product in enumerate(products, start=1):
        product_keys = product_identity_keys(product)
        matches_to_apply: list[tuple[dict[str, Any], str, str, set[str] | None, bool]] = []
        for product_key in product_keys:
            for explicit_match in explicit_matches.get(product_key, []):
                explicit_key = explicit_match.get("typical_key") or ""
                if explicit_key and explicit_key in typical_by_key:
                    selected = explicit_match.get("selected_attributes") or set()
                    selected_filter = selected if explicit_match.get("has_selected_attributes") else None
                    matches_to_apply.append(
                        (
                            typical_by_key[explicit_key],
                            explicit_key,
                            "manual_match",
                            selected_filter,
                            bool(explicit_match.get("apply_type_series_to_all")),
                        )
                    )
        if not matches_to_apply:
            for key in product_keys:
                if key in typical_by_key:
                    matches_to_apply.append((typical_by_key[key], key, "same_identity", None, False))
                    break
        if not matches_to_apply:
            continue
        attrs = product_attributes(product)
        existing = {attr_identity(attr) for attr in attrs if attr_has_import_value(attr)}
        variant_row_indices = type_series_row_indices(attrs)
        for matching_typical, matching_key, match_source, selected_attributes, apply_type_series_to_all in matches_to_apply:
            for typical_attr in product_attributes(matching_typical):
                if not attr_has_import_value(typical_attr):
                    continue
                identity = attr_identity(typical_attr)
                identity_key = f"{identity[0]}:{identity[1]}:{identity[2]}"
                if selected_attributes is not None and identity_key not in selected_attributes:
                    continue
                copied = typical_attribute_target(typical_attr, enrichment_session)
                if copied is None:
                    continue
                copies = [copied]
                if apply_type_series_to_all and attr_is_type_series_row(copied) and variant_row_indices:
                    copies = []
                    for row_i in variant_row_indices:
                        row_copy = dict(copied)
                        row_copy["RowI"] = row_i
                        copies.append(row_copy)
                for copy in copies:
                    identity = attr_identity(copy)
                    if identity in existing:
                        continue
                    attrs.append(copy)
                    existing.add(identity)
                    applied.append(
                        {
                            "source": "typical",
                            "product_index": index,
                            "product_key": matching_key,
                            "match_source": match_source,
                            "attribute_id": copy.get("AttributeId"),
                            "parent_attribute_id": copy.get("ParentAttributeId") or 0,
                            "row_i": copy.get("RowI") or 0,
                        }
                    )
    return {
        "typical_products_count": len(typical_products),
        "matched_products_count": len({item["product_index"] for item in applied}),
        "applied_count": len(applied),
        "applied": applied,
    }


def convert_products_file(
    filename: str,
    content: bytes,
    output_root: Path,
    *,
    product_mapping: dict[str, str] | None = None,
    product_mapping_profile: dict[str, Any] | None = None,
    enrichment_session: dict[str, Any] | None = None,
    typical_products_payload: Any = None,
    product_model_files: dict[str, bytes | str] | None = None,
    product_root_model_id: int | None = None,
) -> dict[str, Any]:
    tables = read_source_tables(filename, content)
    product_root_model_id = int_value(product_root_model_id)
    product_rows = choose_product_rows(tables)
    if product_mapping_profile:
        product_rows = apply_mapping_profile_to_rows(product_rows, product_mapping_profile)
    elif product_mapping:
        product_rows = [apply_column_mapping(row, product_mapping) for row in product_rows]
    enrichment_report = apply_enrichment_session_to_rows(product_rows, enrichment_session)
    mapped_rows = unique_mapped_products([map_source_row(row) for row in product_rows])
    export_schema = export_schema_from_pim_bundle(product_model_files, root_model_id=product_root_model_id)
    products = [build_pim_product(mapped, index, export_schema=export_schema) for index, mapped in enumerate(mapped_rows, start=1)]
    typical_enrichment_report = apply_typical_products_to_products(products, typical_products_payload, enrichment_session)
    report = build_mapping_report(
        filename,
        tables,
        product_rows,
        mapped_rows,
        products,
        product_mapping=product_mapping,
        product_mapping_profile=product_mapping_profile,
    )
    if enrichment_report:
        report["enrichment"] = enrichment_report
    if typical_enrichment_report:
        report["typical_enrichment"] = typical_enrichment_report

    job_id = make_job_id(f"products:{filename}", content)
    output_dir = output_root / job_id
    output_dir.mkdir(parents=True, exist_ok=True)

    products_payload = {
        "productsCount": len(products),
        "products": products,
    }
    (output_dir / "products.json").write_text(
        json.dumps(products_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (output_dir / "mapping_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (output_dir / "mapping_report.xlsx").write_bytes(mapping_report_xlsx_bytes(report))
    (output_dir / "products_acceptance.xlsx").write_bytes(
        product_acceptance_xlsx_bytes(
            products_payload,
            source_file=filename,
            attribute_labels=product_acceptance_attribute_labels(
                export_schema,
                product_mapping_profile,
                product_model_files,
            ),
        )
    )
    if enrichment_session:
        (output_dir / "enrichment_session.json").write_text(
            json.dumps(
                {
                    "source_file": filename,
                    "session": enrichment_session,
                    "applied": {
                        "manual": enrichment_report,
                        "typical": typical_enrichment_report,
                    },
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

    return {
        "job_id": job_id,
        "status": "products_converted",
        "products_count": len(products),
        "output_dir": str(output_dir),
        "files": {
            "products_json": f"/outputs/{job_id}/products.json",
            "products_acceptance_xlsx": f"/outputs/{job_id}/products_acceptance.xlsx",
            "mapping_report_json": f"/outputs/{job_id}/mapping_report.json",
            "mapping_report_xlsx": f"/outputs/{job_id}/mapping_report.xlsx",
            **({"enrichment_session_json": f"/outputs/{job_id}/enrichment_session.json"} if enrichment_session else {}),
        },
        "report": report,
    }


def product_acceptance_attribute_labels(
    export_schema: PimExportSchema,
    product_mapping_profile: dict[str, Any] | None = None,
    product_model_files: dict[str, bytes | str] | None = None,
) -> dict[int, str]:
    labels: dict[int, str] = {
        PIM_ATTR["product_name"]: "Nazwa produktu",
        PIM_ATTR["external_id"]: "Kod produktu",
        PIM_ATTR["unit"]: "Jednostka",
        PIM_ATTR["categories"]: "Kategoria",
        PIM_ATTR["available"]: "Dostępny",
        PIM_ATTR["new"]: "Nowość",
    }
    if product_model_files:
        keyed_files = {pim_bundle_file_key(filename): content for filename, content in product_model_files.items()}
        attributes_content = keyed_files.get("productsattributes")
        if attributes_content is not None:
            try:
                attributes_payload = load_json_content(attributes_content)
                for attribute in pim_items(attributes_payload, "attributes"):
                    attribute_id = int_value(attribute.get("Id"))
                    if attribute_id is None or is_deleted(attribute):
                        continue
                    label = (
                        attribute.get("DispName")
                        or attribute.get("AttributeName")
                        or attribute.get("Name")
                        or attribute.get("Code")
                    )
                    if label not in (None, ""):
                        unit = attribute.get("Unit")
                        labels[attribute_id] = f"{label} [{unit}]" if unit not in (None, "") else str(label)
            except (json.JSONDecodeError, UnicodeDecodeError, TypeError):
                pass

    for item in (product_mapping_profile or {}).values():
        if not isinstance(item, dict):
            continue
        target_path = str(item.get("target_path") or "")
        attribute_id = export_schema.product_attribute_id(target_path) or export_schema.type_series_attribute_id(target_path)
        if attribute_id is None:
            continue
        target_label = item.get("target_label") or item.get("label")
        if target_label not in (None, ""):
            labels[attribute_id] = str(target_label)
    return labels


def convert_systems_file(
    filename: str,
    content: bytes,
    products_content: bytes,
    output_root: Path,
    *,
    system_mapping: dict[str, str] | None = None,
) -> dict[str, Any]:
    tables = read_source_tables(filename, content)
    system_rows = choose_system_rows(tables)
    if system_mapping:
        system_rows = [apply_column_mapping(row, system_mapping) for row in system_rows]

    products_payload = json.loads(products_content.decode("utf-8-sig"))
    product_index = build_product_index(products_payload.get("products", []))
    system_mapped_rows = [map_system_row(row) for row in system_rows]
    building_elements, system_warnings = build_building_elements(system_mapped_rows, product_index)
    report = build_mapping_report(
        filename,
        tables,
        [],
        [],
        products_payload.get("products", []),
        system_rows=system_rows,
        system_mapped_rows=system_mapped_rows,
        building_elements=building_elements,
        system_warnings=system_warnings,
    )

    job_id = stable_hash(f"systems:{filename}", content[:2048], len(content), products_content[:2048], len(products_content))[:12]
    output_dir = output_root / job_id
    output_dir.mkdir(parents=True, exist_ok=True)

    building_elements_payload = {
        "buildingElementsCount": len(building_elements),
        "buildingElements": building_elements,
    }
    (output_dir / "building_elements.json").write_text(
        json.dumps(building_elements_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (output_dir / "mapping_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return {
        "job_id": job_id,
        "status": "systems_converted",
        "products_count": products_payload.get("productsCount", len(products_payload.get("products", []))),
        "building_elements_count": len(building_elements),
        "output_dir": str(output_dir),
        "files": {
            "building_elements_json": f"/outputs/{job_id}/building_elements.json",
            "mapping_report_json": f"/outputs/{job_id}/mapping_report.json",
        },
        "report": report,
    }


def read_source_tables(filename: str, content: bytes) -> list[SourceTable]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".json":
        return read_json_tables(content)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_tables(content)
    if suffix in {".csv", ".tsv"}:
        return [SourceTable(Path(filename).stem, read_delimited_rows(content, suffix))]
    raise ValueError(f"Unsupported file type: {suffix}. Use .xlsx, .json, .csv or .tsv.")


def read_json_tables(content: bytes) -> list[SourceTable]:
    data = json.loads(content.decode("utf-8-sig"))
    if isinstance(data, list):
        return [SourceTable("json", [normalize_json_row(row) for row in data if isinstance(row, dict)])]
    if isinstance(data, dict):
        tables: list[SourceTable] = []
        for key, value in data.items():
            if isinstance(value, list):
                rows = [normalize_json_row(row) for row in value if isinstance(row, dict)]
                if rows:
                    tables.append(SourceTable(str(key), rows))
        if tables:
            return tables
        return [SourceTable("json", [normalize_json_row(data)])]
    raise ValueError("JSON must be an object or a list of objects.")


def normalize_json_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = {key: normalize_json_cell(value) for key, value in row.items()}
    add_file_helper_columns(normalized)
    return normalized


def normalize_json_cell(value: Any) -> Any:
    if isinstance(value, dict) and "value" in value and set(value.keys()) <= {"label", "value"}:
        return value.get("value")
    return value


def add_file_helper_columns(row: dict[str, Any]) -> None:
    files = file_items(row.get("files"))
    if not files:
        return
    image_urls = [str(item.get("url")) for item in files if is_image_file(item) and item.get("url")]
    document_urls = [str(item.get("url")) for item in files if not is_image_file(item) and item.get("url")]
    if image_urls and "packshot_url" not in row:
        row["packshot_url"] = image_urls[0]
    if image_urls and "image_urls" not in row:
        row["image_urls"] = "\n".join(image_urls)
    if document_urls and "document_urls" not in row:
        row["document_urls"] = "\n".join(document_urls)


def read_xlsx_tables(content: bytes) -> list[SourceTable]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    tables: list[SourceTable] = []
    for sheet in workbook.worksheets:
        values = list(sheet.iter_rows(values_only=True))
        rows = rows_from_matrix(values)
        if rows:
            tables.append(SourceTable(sheet.title, rows))
    if not tables:
        raise ValueError("No tabular rows detected in workbook.")
    return tables


def decode_delimited_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1250", "iso-8859-2", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8-sig", errors="replace")


def detect_delimiter(text: str, suffix: str) -> str:
    if suffix == ".tsv":
        return "\t"
    sample = text[:4096]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
    except csv.Error:
        return ";" if sample.count(";") > sample.count(",") else ","


def read_delimited_rows(content: bytes, suffix: str) -> list[dict[str, Any]]:
    text = decode_delimited_text(content)
    delimiter = detect_delimiter(text, suffix)
    reader = csv.DictReader(StringIO(text), delimiter=delimiter)
    return [{key: value for key, value in row.items()} for row in reader]


def rows_from_matrix(values: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    header_index = None
    for index, row in enumerate(values[:25]):
        non_empty = [cell for cell in row if cell not in (None, "")]
        if len(non_empty) >= 2:
            header_index = index
            break
    if header_index is None:
        return []

    headers = [normalize_header(cell) or f"column_{i + 1}" for i, cell in enumerate(values[header_index])]
    rows: list[dict[str, Any]] = []
    for raw_row in values[header_index + 1 :]:
        row = {headers[i]: value for i, value in enumerate(raw_row) if i < len(headers)}
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows


def choose_product_rows(tables: list[SourceTable]) -> list[dict[str, Any]]:
    scored = sorted(tables, key=lambda table: table_score(table), reverse=True)
    if not scored or table_score(scored[0]) == 0:
        raise ValueError("No product-like table detected. Add columns such as name/nazwa, SKU/kod or description/opis.")
    return scored[0].rows


def choose_system_rows(tables: list[SourceTable]) -> list[dict[str, Any]]:
    scored = sorted(tables, key=lambda table: system_table_score(table), reverse=True)
    if not scored or system_table_score(scored[0]) == 0:
        return []
    return scored[0].rows


def table_score(table: SourceTable) -> int:
    if not table.rows:
        return 0
    headers = set(table.rows[0].keys())
    canonical = {guess_field(header) for header in headers}
    score = len([field for field in canonical if field])
    if "product_name" in canonical:
        score += 4
    if "external_id" in canonical:
        score += 2
    return score


def system_table_score(table: SourceTable) -> int:
    if not table.rows:
        return 0
    headers = set(table.rows[0].keys())
    canonical = {guess_system_field(header) for header in headers}
    score = len([field for field in canonical if field])
    if "system_name" in canonical:
        score += 4
    if "variant_name" in canonical:
        score += 2
    if "layer_name" in canonical:
        score += 2
    if "product_code" in canonical or "product_name" in canonical:
        score += 2
    return score if {"system_name", "variant_name", "layer_name"} & canonical else 0


def map_source_row(row: dict[str, Any]) -> dict[str, Any]:
    mapped: dict[str, Any] = {
        "core": {},
        "product_information": {},
        "package": {},
        "sot": {},
        "documents": [],
        "pim_attributes": [],
        "unmapped": {},
        "column_mapping": {},
    }
    for raw_key, value in row.items():
        if value in (None, ""):
            continue
        field = pim_export_field_from_canonical(raw_key) or pim_attribute_field(raw_key) or guess_field(raw_key)
        mapped["column_mapping"][raw_key] = field or "unmapped"
        if field in {"product_name", "external_id", "unit", "categories"}:
            mapped["core"][field] = value
        elif field in PRODUCT_INFO_ATTR:
            mapped["product_information"][field] = value
        elif field in PACKAGE_ATTR:
            mapped["package"][field] = value
        elif field in SOT_ATTR:
            mapped["sot"][field] = value
        elif raw_key == "_type_series_rows" and isinstance(value, list):
            mapped["type_series_rows"] = value
        elif field == "documents":
            mapped["documents"].extend(document_links_from_value(value))
        elif pim_attribute_id := pim_attribute_id_from_field(field):
            mapped["pim_attributes"].append({"attribute_id": pim_attribute_id, "value": value})
        else:
            mapped["unmapped"][raw_key] = value

    if "product_name" not in mapped["core"]:
        fallback = mapped["core"].get("external_id") or mapped["product_information"].get("full_name")
        mapped["core"]["product_name"] = str(fallback or "Produkt bez nazwy")
    return mapped


def pim_export_field_from_canonical(field: Any) -> str | None:
    canonical = str(field)
    return {
        "product.name.value": "product_name",
        "product.code.value": "external_id",
        "product.unit.value": "unit",
        "product.category[].value": "categories",
        "product.description.value": "description",
        "product.properties.value": "properties",
        "product.application.value": "prosperation",
        "product.usage_method.value": "usage_method",
        "product.norms.value": "norms",
        "product.manufacturer.value": "manufacturer",
        "product.product_url.value": "url",
        "product.documents[].url.value": "documents",
        "product.documents[].name.value": "documents",
        "product.documents[].category.value": "documents",
        "product.documents[].extension.value": "documents",
        "product.documents[].safe_name.value": "documents",
        "product.packages[].raw_text.value": "package_text",
        "product.packages[].weight.value": "weight",
        "product.packages[].capacity.value": "capacity",
        "type_series[].thickness.value": "thickness",
        "type_series[].lambda_value.value": "lambda",
        "type_series[].density.value": "density",
        "type_series[].vapor_permeability_mu.value": "vapor_permeability",
    }.get(canonical)


def pim_attribute_id_from_field(field: Any) -> int | None:
    match = re.fullmatch(r"pim\.attribute\.(\d+)\.value", str(field or ""))
    return int(match.group(1)) if match else None


def pim_attribute_field(field: Any) -> str | None:
    return str(field) if pim_attribute_id_from_field(field) else None


def document_links_from_value(value: Any) -> list[dict[str, Any]]:
    files = file_items(value)
    if files:
        return [document_link_from_file(item) for item in files if item.get("url")]
    links = links_from_text(value)
    return [document_link_from_url(url) for url in links]


def file_items(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    if isinstance(value, dict):
        if "url" in value:
            return [value]
        if "value" in value:
            return file_items(value.get("value"))
        return []
    if isinstance(value, str):
        text = value.strip()
        if text.startswith("[") and text.endswith("]"):
            try:
                parsed = json.loads(text)
            except json.JSONDecodeError:
                return []
            return file_items(parsed)
    return []


def links_from_text(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, (list, tuple, set)):
        result: list[str] = []
        for item in value:
            result.extend(links_from_text(item))
        return result
    text = str(value)
    return re.findall(r"https?://[^\s,;|]+", text)


def document_link_from_file(item: dict[str, Any]) -> dict[str, Any]:
    url = str(item.get("url") or "")
    name = str(item.get("caption") or file_name_from_url(url) or item.get("id") or "file")
    category = str(item.get("type") or category_from_url(url) or "file")
    return {
        "name": name,
        "url": url,
        "category": category,
        "extension": file_extension(name) or file_extension(url),
        "safe_name": safe_file_name(name),
        "display_web": True,
    }


def document_link_from_url(url: str) -> dict[str, Any]:
    name = file_name_from_url(url) or url
    return {
        "name": name,
        "url": url,
        "category": category_from_url(url),
        "extension": file_extension(url),
        "safe_name": safe_file_name(name),
        "display_web": True,
    }


def is_image_file(item: dict[str, Any]) -> bool:
    file_type = normalize_header(item.get("type"))
    url = str(item.get("url") or item.get("caption") or "")
    return file_type == "image" or file_extension(url) in {"jpg", "jpeg", "png", "webp", "gif"}


def category_from_url(url: str) -> str:
    extension = file_extension(url)
    if extension in {"jpg", "jpeg", "png", "webp", "gif"}:
        return "image"
    if extension == "pdf":
        return "pdf"
    return "file"


def file_name_from_url(url: str) -> str:
    path = urlparse(str(url or "")).path
    name = unquote(Path(path).name)
    return name


def file_extension(value: Any) -> str:
    suffix = Path(urlparse(str(value or "")).path).suffix.lower().lstrip(".")
    return suffix


def safe_file_name(value: Any) -> str:
    name = file_name_from_url(str(value)) or str(value or "")
    name = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_")
    return name[:180] or "file"


def unique_mapped_products(mapped_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    seen: set[str] = set()
    for mapped in mapped_rows:
        core = mapped["core"]
        key = normalize_lookup(core.get("external_id") or core.get("product_name"))
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(mapped)
    return result


def map_system_row(row: dict[str, Any]) -> dict[str, Any]:
    mapped: dict[str, Any] = {
        "system": {},
        "variant": {},
        "layer": {},
        "product": {},
        "column_mapping": {},
        "unmapped": {},
    }
    for raw_key, value in row.items():
        if value in (None, ""):
            continue
        field = guess_system_field(raw_key)
        mapped["column_mapping"][raw_key] = field or "unmapped"
        if field in {"system_name", "building_element_type", "insulation_type", "bim_building_element_type"}:
            mapped["system"][field] = value
        elif field == "variant_name":
            mapped["variant"][field] = value
        elif field in {"layer_position", "layer_name"}:
            mapped["layer"][field] = value
        elif field in {"product_code", "product_name", "default", "quantity"}:
            mapped["product"][field] = value
        else:
            mapped["unmapped"][raw_key] = value
    return mapped


def build_pim_product(
    mapped: dict[str, Any],
    index: int,
    *,
    export_schema: PimExportSchema = DEFAULT_EXPORT_SCHEMA,
) -> dict[str, Any]:
    product_id = PRODUCT_ID_START + index
    attrs: list[dict[str, Any]] = []
    core = mapped["core"]

    product_name_attr = export_schema.product_attribute_id("product.name.value", PIM_ATTR["product_name"])
    if product_name_attr is not None:
        add_attr(
            attrs,
            product_name_attr,
            varchar=str(core["product_name"]),
            parent_attribute_id=export_schema.product_parent_for_attribute(product_name_attr),
        )
    product_code_attr = export_schema.product_attribute_id("product.code.value", PIM_ATTR["external_id"])
    if product_code_attr is not None:
        add_attr(
            attrs,
            product_code_attr,
            varchar=str(core.get("external_id") or generated_external_id(core["product_name"], product_id)),
            parent_attribute_id=export_schema.product_parent_for_attribute(product_code_attr),
        )
    unit_attr = export_schema.product_attribute_id("product.unit.value", PIM_ATTR["unit"])
    if core.get("unit") and unit_attr is not None:
        add_attr(attrs, unit_attr, varchar=str(core["unit"]), parent_attribute_id=export_schema.product_parent_for_attribute(unit_attr))
    available_attr = export_schema.static_attribute_id(PIM_ATTR["available"])
    if available_attr is not None:
        add_attr(attrs, available_attr, boolean=True, parent_attribute_id=export_schema.product_parent_for_attribute(available_attr))
    new_attr = export_schema.static_attribute_id(PIM_ATTR["new"])
    if new_attr is not None:
        add_attr(attrs, new_attr, boolean=True, parent_attribute_id=export_schema.product_parent_for_attribute(new_attr))

    category_attr = export_schema.product_attribute_id("product.category[].value", PIM_ATTR["categories"])
    for category_id in category_ids(core.get("categories")):
        if category_attr is not None:
            add_attr(
                attrs,
                category_attr,
                int_value=category_id,
                boolean=True,
                parent_attribute_id=export_schema.product_parent_for_attribute(category_attr),
            )

    add_product_information(attrs, mapped["product_information"], export_schema)
    add_package(attrs, mapped["package"])
    add_sot(attrs, mapped["sot"], export_schema)
    add_sot_rows(attrs, mapped.get("type_series_rows") or [], export_schema)
    add_documents(attrs, mapped["documents"])
    add_generic_pim_attributes(attrs, mapped["pim_attributes"], export_schema)

    return {
        "Id": product_id,
        "prodctTypeId": PRODUCT_TYPE_ID,
        "ModelType": export_schema.product_model_id,
        "dataVersions": [
            {
                "VersionId": 1,
                "productAttributes": attrs,
            }
        ],
    }


def build_product_index(products: list[dict[str, Any]]) -> dict[str, int]:
    index: dict[str, int] = {}
    for product in products:
        attrs = product["dataVersions"][0]["productAttributes"]
        product_id = product["Id"]
        for attr in attrs:
            if attr["AttributeId"] in {PIM_ATTR["product_name"], PIM_ATTR["external_id"]}:
                value = attr.get("varcharValue") or attr.get("TextValue")
                if value:
                    index[normalize_lookup(value)] = product_id
    return index


def build_building_elements(
    rows: list[dict[str, Any]],
    product_index: dict[str, int],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not rows:
        return [], {"systems_table_missing": True, "unmatched_products": []}

    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        system_name = str(row["system"].get("system_name") or "System bez nazwy")
        grouped.setdefault(system_name, []).append(row)

    warnings = {"systems_table_missing": False, "unmatched_products": []}
    elements = []
    for index, (system_name, system_rows) in enumerate(grouped.items(), start=1):
        elements.append(build_building_element(system_name, system_rows, index, product_index, warnings))
    return elements, warnings


def build_building_element(
    system_name: str,
    rows: list[dict[str, Any]],
    index: int,
    product_index: dict[str, int],
    warnings: dict[str, Any],
) -> dict[str, Any]:
    attrs: list[dict[str, Any]] = []
    first_system = rows[0]["system"]
    add_attr(attrs, BUILDING_ELEMENT_ATTR["name"], varchar=system_name)
    add_select_attr(attrs, BUILDING_ELEMENT_ATTR["building_element_type"], first_system.get("building_element_type"), BUILDING_ELEMENT_TYPE_OPTIONS)
    add_select_attr(attrs, BUILDING_ELEMENT_ATTR["insulation_type"], first_system.get("insulation_type"), INSULATION_TYPE_OPTIONS)
    add_select_attr(attrs, BUILDING_ELEMENT_ATTR["bim_building_element_type"], first_system.get("bim_building_element_type"), BIM_TYPE_OPTIONS)

    variant_groups: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        variant_name = str(row["variant"].get("variant_name") or "Wariant podstawowy")
        variant_groups.setdefault(variant_name, []).append(row)

    for variant_i, (variant_name, variant_rows) in enumerate(variant_groups.items()):
        variant_hash = stable_hash(system_name, variant_name)
        add_attr(
            attrs,
            BUILDING_ELEMENT_ATTR["variant_name"],
            varchar=variant_name,
            parent_attribute_id=BUILDING_ELEMENT_ATTR["variants"],
            row_hash=variant_hash,
            parent_hash=None,
            row_i=variant_i,
        )
        layer_groups: dict[tuple[int, str], list[dict[str, Any]]] = {}
        for row in variant_rows:
            position = parse_layer_position(row["layer"].get("layer_position"), len(layer_groups) + 1)
            layer_name = str(row["layer"].get("layer_name") or f"Layer {position}")
            layer_groups.setdefault((position, layer_name), []).append(row)

        for layer_i, ((position, layer_name), layer_rows) in enumerate(sorted(layer_groups.items())):
            layer_hash = stable_hash(system_name, variant_name, position, layer_name)
            add_attr(
                attrs,
                BUILDING_ELEMENT_ATTR["layer_position"],
                int_value=LAYER_POSITION_OPTIONS.get(position, LAYER_POSITION_OPTIONS[1]),
                parent_attribute_id=BUILDING_ELEMENT_ATTR["layers"],
                row_hash=layer_hash,
                parent_hash=variant_hash,
                row_i=layer_i,
            )
            add_attr(
                attrs,
                BUILDING_ELEMENT_ATTR["layer_name"],
                varchar=layer_name,
                parent_attribute_id=BUILDING_ELEMENT_ATTR["layers"],
                row_hash=layer_hash,
                parent_hash=variant_hash,
                row_i=layer_i,
            )
            for product_i, row in enumerate(layer_rows):
                product_id = resolve_product_id(row["product"], product_index)
                product_key = row["product"].get("product_code") or row["product"].get("product_name")
                if product_id is None:
                    warnings["unmatched_products"].append(
                        {
                            "system_name": system_name,
                            "variant_name": variant_name,
                            "layer_name": layer_name,
                            "product": product_key,
                        }
                    )
                    continue
                available_hash = stable_hash(system_name, variant_name, layer_name, product_id, product_i)
                add_attr(
                    attrs,
                    BUILDING_ELEMENT_ATTR["product"],
                    int_value=product_id,
                    int_value2=parse_int(row["product"].get("quantity")),
                    parent_attribute_id=BUILDING_ELEMENT_ATTR["available_products"],
                    row_hash=available_hash,
                    parent_hash=layer_hash,
                    row_i=product_i,
                )
                add_attr(
                    attrs,
                    BUILDING_ELEMENT_ATTR["default"],
                    boolean=parse_bool(row["product"].get("default")),
                    parent_attribute_id=BUILDING_ELEMENT_ATTR["available_products"],
                    row_hash=available_hash,
                    parent_hash=layer_hash,
                    row_i=product_i,
                )

    return {
        "Id": BUILDING_ELEMENT_ID_START + index,
        "elementTypeId": BUILDING_ELEMENT_TYPE_ID,
        "ModelType": BUILDING_ELEMENT_MODEL_TYPE,
        "dataVersions": [
            {
                "VersionId": 1,
                "productAttributes": attrs,
                "filesAttributes": [],
                "colorsAttributes": [],
            }
        ],
    }


def add_product_information(
    attrs: list[dict[str, Any]],
    info: dict[str, Any],
    export_schema: PimExportSchema = DEFAULT_EXPORT_SCHEMA,
) -> None:
    if not info:
        return
    row_hash = stable_hash("product_information", info)
    for field, value in info.items():
        target_path = PRODUCT_INFO_FIELD_TARGETS.get(field)
        attr_id = export_schema.product_attribute_id(target_path or "", PRODUCT_INFO_ATTR.get(field))
        if attr_id:
            add_attr(
                attrs,
                attr_id,
                text=str(value),
                parent_attribute_id=export_schema.product_parent_for_attribute(
                    attr_id,
                    fallback=PIM_ATTR["product_information"],
                ),
                row_hash=row_hash if export_schema.product_parent_for_attribute(attr_id, fallback=PIM_ATTR["product_information"]) else None,
            )


def add_package(attrs: list[dict[str, Any]], package: dict[str, Any]) -> None:
    if not package:
        return
    row_hash = stable_hash("package", package)
    for field, value in package.items():
        attr_id = PACKAGE_ATTR.get(field)
        if not attr_id:
            continue
        kwargs = {"number": parse_number(value)} if field in {"weight", "capacity"} else {"varchar": str(value)}
        add_attr(attrs, attr_id, parent_attribute_id=PIM_ATTR["packages"], row_hash=row_hash, **kwargs)


def add_sot(attrs: list[dict[str, Any]], sot: dict[str, Any], export_schema: PimExportSchema = DEFAULT_EXPORT_SCHEMA) -> None:
    if not sot:
        return
    row_hash = stable_hash("sot", sot)
    for field, value in sot.items():
        target_path = SOT_FIELD_TARGETS.get(field)
        attr_id = export_schema.type_series_attribute_id(target_path or "") or SOT_ATTR.get(field)
        number = parse_number(value)
        if attr_id and number is not None:
            add_attr(
                attrs,
                attr_id,
                number=number,
                parent_attribute_id=export_schema.type_series_parent_for_attribute(attr_id),
                row_hash=row_hash,
            )


def add_typed_attr_value(
    attrs: list[dict[str, Any]],
    attribute_id: int,
    value: Any,
    *,
    value_kind: str,
    parent_attribute_id: int = 0,
    row_hash: str | None = None,
    row_i: int = 0,
) -> None:
    if value_kind == "number":
        number = parse_number(value)
        if number is not None:
            add_attr(attrs, attribute_id, number=number, parent_attribute_id=parent_attribute_id, row_hash=row_hash, row_i=row_i)
        return
    if isinstance(value, bool):
        add_attr(attrs, attribute_id, boolean=value, parent_attribute_id=parent_attribute_id, row_hash=row_hash, row_i=row_i)
        return
    text = str(value)
    kwargs = {"varchar": text} if len(text) <= 255 else {"text": text}
    add_attr(attrs, attribute_id, parent_attribute_id=parent_attribute_id, row_hash=row_hash, row_i=row_i, **kwargs)


def add_sot_rows(
    attrs: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    export_schema: PimExportSchema = DEFAULT_EXPORT_SCHEMA,
) -> None:
    for index, row in enumerate(rows):
        if not row:
            continue
        row_hash = stable_hash("sot", row, str(index))
        for key, value in row.items():
            if value in (None, ""):
                continue
            canonical_field = pim_export_field_from_canonical(key) or ""
            canonical_target = SOT_FIELD_TARGETS.get(canonical_field)
            attr_id = export_schema.type_series_attribute_id(str(key))
            if not attr_id and canonical_target:
                attr_id = export_schema.type_series_attribute_id(canonical_target)
            if not attr_id:
                attr_id = SOT_ATTR.get(canonical_field)
            if attr_id:
                add_typed_attr_value(
                    attrs,
                    attr_id,
                    value,
                    value_kind=export_schema.attribute_value_kind(attr_id, "number"),
                    parent_attribute_id=export_schema.type_series_parent_for_attribute(attr_id),
                    row_hash=row_hash,
                    row_i=index,
                )
                continue
            pim_attr_id = pim_attribute_id_from_field(key)
            if pim_attr_id:
                add_typed_attr_value(
                    attrs,
                    pim_attr_id,
                    value,
                    value_kind=export_schema.attribute_value_kind(pim_attr_id),
                    parent_attribute_id=export_schema.type_series_parent_for_attribute(pim_attr_id),
                    row_hash=row_hash,
                    row_i=index,
                )


def add_documents(attrs: list[dict[str, Any]], documents: list[dict[str, Any]]) -> None:
    seen_urls: set[str] = set()
    for index, document in enumerate(documents):
        url = str(document.get("url") or "")
        if not url or url in seen_urls:
            continue
        seen_urls.add(url)
        row_hash = stable_hash("document", document, str(index))
        add_attr(attrs, DOCUMENT_ATTR["name"], varchar=str(document.get("name") or "file"), parent_attribute_id=PIM_ATTR["documents"], row_hash=row_hash, row_i=index)
        add_attr(attrs, DOCUMENT_ATTR["url"], varchar=url, parent_attribute_id=PIM_ATTR["documents"], row_hash=row_hash, row_i=index)
        if document.get("category"):
            add_attr(attrs, DOCUMENT_ATTR["category"], varchar=str(document["category"]), parent_attribute_id=PIM_ATTR["documents"], row_hash=row_hash, row_i=index)
        if document.get("extension"):
            add_attr(attrs, DOCUMENT_ATTR["extension"], varchar=str(document["extension"]), parent_attribute_id=PIM_ATTR["documents"], row_hash=row_hash, row_i=index)
        if document.get("safe_name"):
            add_attr(attrs, DOCUMENT_ATTR["safe_name"], varchar=str(document["safe_name"]), parent_attribute_id=PIM_ATTR["documents"], row_hash=row_hash, row_i=index)
        add_attr(attrs, DOCUMENT_ATTR["display_web"], boolean=bool(document.get("display_web", True)), parent_attribute_id=PIM_ATTR["documents"], row_hash=row_hash, row_i=index)


def add_generic_pim_attributes(
    attrs: list[dict[str, Any]],
    pim_attributes: list[dict[str, Any]],
    export_schema: PimExportSchema = DEFAULT_EXPORT_SCHEMA,
) -> None:
    for item in pim_attributes:
        attribute_id = parse_int(item.get("attribute_id"))
        value = item.get("value")
        if attribute_id is None or value in (None, ""):
            continue
        if not export_schema.has_attribute(attribute_id):
            continue
        parent_attribute_id = export_schema.product_parent_for_attribute(attribute_id, fallback=GENERIC_PIM_PARENT_IDS.get(attribute_id, 0))
        row_hash = stable_hash("pim", parent_attribute_id) if parent_attribute_id else None
        if isinstance(value, list):
            for option in value:
                add_pim_option_or_value(attrs, attribute_id, option, parent_attribute_id=parent_attribute_id, row_hash=row_hash)
            continue
        if isinstance(value, dict):
            add_pim_option_or_value(attrs, attribute_id, value, parent_attribute_id=parent_attribute_id, row_hash=row_hash)
            continue
        kwargs: dict[str, Any]
        if isinstance(value, bool):
            kwargs = {"boolean": value}
        else:
            text = str(value)
            kwargs = {"varchar": text} if len(text) <= 255 else {"text": text}
        add_attr(attrs, attribute_id, parent_attribute_id=parent_attribute_id, row_hash=row_hash, **kwargs)


def add_pim_option_or_value(
    attrs: list[dict[str, Any]],
    attribute_id: int,
    option: Any,
    *,
    parent_attribute_id: int = 0,
    row_hash: str | None = None,
) -> None:
    if isinstance(option, dict):
        if option.get("unmatched"):
            return
        option_id = parse_int(option.get("id"))
        if option_id is not None:
            add_attr(attrs, attribute_id, int_value=option_id, boolean=True, parent_attribute_id=parent_attribute_id, row_hash=row_hash)
            return
        value = option.get("raw") or option.get("label") or option.get("value")
    else:
        value = option
    if value not in (None, ""):
        text = str(value)
        kwargs = {"varchar": text} if len(text) <= 255 else {"text": text}
        add_attr(attrs, attribute_id, parent_attribute_id=parent_attribute_id, row_hash=row_hash, **kwargs)


def add_attr(
    attrs: list[dict[str, Any]],
    attribute_id: int,
    *,
    varchar: str | None = None,
    text: str | None = None,
    int_value: int | None = None,
    number: float | None = None,
    boolean: bool = False,
    int_value2: int | None = None,
    parent_attribute_id: int = 0,
    row_hash: str | None = None,
    parent_hash: str | None = "",
    row_i: int = 0,
) -> None:
    attrs.append(
        {
            "AttributeId": attribute_id,
            "hash": row_hash,
            "parentHash": parent_hash,
            "ParentAttributeId": parent_attribute_id,
            "varcharValue": varchar,
            "TextValue": text,
            "IntValue": int_value,
            "IntValue2": int_value2,
            "NumberValue": number,
            "BooleanValue": boolean,
            "MainAttributeId": None,
            "RowI": row_i,
        }
    )


def add_select_attr(attrs: list[dict[str, Any]], attribute_id: int, value: Any, options: dict[str, int]) -> None:
    if value in (None, ""):
        return
    option_id = options.get(normalize_header(value))
    if option_id is not None:
        add_attr(attrs, attribute_id, int_value=option_id)


def build_mapping_report(
    filename: str,
    tables: list[SourceTable],
    rows: list[dict[str, Any]],
    mapped_rows: list[dict[str, Any]],
    products: list[dict[str, Any]],
    *,
    system_rows: list[dict[str, Any]] | None = None,
    system_mapped_rows: list[dict[str, Any]] | None = None,
    building_elements: list[dict[str, Any]] | None = None,
    system_warnings: dict[str, Any] | None = None,
    product_mapping: dict[str, str] | None = None,
    product_mapping_profile: dict[str, Any] | None = None,
) -> dict[str, Any]:
    column_mapping: dict[str, str] = {}
    unmapped_columns: set[str] = set()
    for mapped in mapped_rows:
        column_mapping.update(mapped["column_mapping"])
        unmapped_columns.update(mapped["unmapped"].keys())

    missing_names = [
        index + 1
        for index, mapped in enumerate(mapped_rows)
        if mapped["core"].get("product_name") == "Produkt bez nazwy"
    ]
    categories_unmatched = sorted(
        {
            str(mapped["core"].get("categories"))
            for mapped in mapped_rows
            if mapped["core"].get("categories") and not category_ids(mapped["core"].get("categories"))
        }
    )

    system_column_mapping: dict[str, str] = {}
    system_unmapped_columns: set[str] = set()
    for mapped in system_mapped_rows or []:
        system_column_mapping.update(mapped["column_mapping"])
        system_unmapped_columns.update(mapped["unmapped"].keys())

    return {
        "source_file": filename,
        "tables_detected": [
            {
                "name": table.name,
                "rows": len(table.rows),
                "product_score": table_score(table),
                "system_score": system_table_score(table),
            }
            for table in tables
        ],
        "selected_rows": len(rows),
        "products_generated": len(products),
        "system_rows_selected": len(system_rows or []),
        "building_elements_generated": len(building_elements or []),
        "column_mapping": column_mapping,
        "product_mapping": product_mapping or {},
        "product_mapping_profile": product_mapping_profile or {},
        "system_column_mapping": system_column_mapping,
        "unmapped_columns": sorted(unmapped_columns),
        "system_unmapped_columns": sorted(system_unmapped_columns),
        "warnings": {
            "rows_without_name": missing_names,
            "categories_not_matched_to_pim_options": categories_unmatched,
            **(system_warnings or {}),
        },
        "next_steps": [
            "Review unmapped_columns and add missing PIM model attributes before importing them.",
            "Review unmatched categories and add dictionary mappings where needed.",
            "Review unmatched_products before using building_elements.json in the local PIM agent.",
        ],
    }


def guess_field(raw_header: Any) -> str | None:
    normalized = normalize_header(raw_header)
    for canonical, aliases in CANONICAL_ALIASES.items():
        if normalized in {normalize_header(alias) for alias in aliases}:
            return canonical
    compact = normalized.replace(" ", "")
    if compact in {"lambdaw/mk", "lambdawmk"}:
        return "lambda"
    return None


def guess_system_field(raw_header: Any) -> str | None:
    normalized = normalize_header(raw_header)
    for canonical, aliases in SYSTEM_ALIASES.items():
        if normalized in {normalize_header(alias) for alias in aliases}:
            return canonical
    return None


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"[_\-/]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text


def category_ids(value: Any) -> list[int]:
    if value in (None, ""):
        return []
    result: list[int] = []
    for part in category_parts(value):
        if part in (None, ""):
            continue
        if isinstance(part, int) and part > 0:
            result.append(part)
            continue
        text = str(part).strip()
        if re.fullmatch(r"\d+", text):
            numeric = int(text)
            if numeric > 0:
                result.append(numeric)
            continue
        option_id = CATEGORY_OPTIONS.get(normalize_header(text))
        if option_id:
            result.append(option_id)
    return sorted(set(result))


def export_schema_from_pim_bundle(files: dict[str, bytes | str] | None, root_model_id: int | None = None) -> PimExportSchema:
    root_model_id = int_value(root_model_id)
    if not files:
        return DEFAULT_EXPORT_SCHEMA
    keyed_files = {pim_bundle_file_key(filename): content for filename, content in files.items()}
    models_content = keyed_files.get("productsmodels")
    attributes_content = keyed_files.get("productsattributes")
    if models_content is None or attributes_content is None:
        return DEFAULT_EXPORT_SCHEMA
    try:
        models_payload = load_json_content(models_content)
        attributes_payload = load_json_content(attributes_content)
    except (json.JSONDecodeError, UnicodeDecodeError, TypeError):
        return DEFAULT_EXPORT_SCHEMA

    models = pim_items(models_payload, "models")
    attributes = [attribute for attribute in pim_items(attributes_payload, "attributes") if not is_deleted(attribute)]
    if not models or not attributes:
        return DEFAULT_EXPORT_SCHEMA

    models_by_id = {model_id: model for model in models if (model_id := int_value(model.get("Id"))) is not None}
    product_model_ids = {
        int(model["Id"])
        for model in models
        if str(model.get("Id") or "").strip()
        and normalize_lookup(model.get("modelType") or model.get("ModelType") or model.get("type")) == "product"
    } or {PRODUCT_MODEL_TYPE}
    if root_model_id is not None:
        product_model_ids = {root_model_id}
    product_model_id = sorted(product_model_ids)[0] if product_model_ids else PRODUCT_MODEL_TYPE

    attributes_by_model: dict[int, list[dict[str, Any]]] = {}
    for attribute in attributes:
        model_id = int_value(attribute.get("ProductModelId"))
        if model_id is not None:
            attributes_by_model.setdefault(model_id, []).append(attribute)
    for grouped in attributes_by_model.values():
        grouped.sort(key=attribute_order)

    product_attribute_ids: dict[str, int] = {}
    product_parent_by_attribute_id: dict[int, int] = {}
    known_attribute_ids = frozenset(
        attribute_id
        for attribute in attributes
        if (attribute_id := int_value(attribute.get("Id"))) is not None
    )
    attribute_value_kinds = {
        attribute_id: value_kind_from_attribute(attribute)
        for attribute in attributes
        if (attribute_id := int_value(attribute.get("Id"))) is not None
    }
    parent_id: int | None = None
    attribute_ids: dict[str, int] = {}
    parent_by_attribute_id: dict[int, int] = {}
    for model_id in sorted(product_model_ids):
        for attribute in attributes_by_model.get(model_id, []):
            if skip_pim_attribute(attribute):
                continue
            target_model_id = int_value(attribute.get("TargetModelId"))
            attribute_id = int_value(attribute.get("Id"))
            if not target_model_id or not is_nested_attribute(attribute):
                if attribute_id is None:
                    continue
                target_path = semantic_pim_field_key(attribute) or f"pim.attribute.{attribute_id}.value"
                product_attribute_ids[target_path] = attribute_id
                product_parent_by_attribute_id[attribute_id] = 0
                continue
            target_model = models_by_id.get(target_model_id)
            child_attributes = [
                child
                for child in attributes_by_model.get(target_model_id, [])
                if not skip_pim_attribute(child) and not is_nested_attribute(child)
            ]
            current_parent_id = int_value(attribute.get("Id"))
            if current_parent_id is None:
                continue
            if not is_type_series_parent(attribute, target_model, child_attributes):
                for child in child_attributes:
                    child_id = int_value(child.get("Id"))
                    if child_id is None:
                        continue
                    target_path = semantic_pim_field_key(
                        child,
                        parent_attribute=attribute,
                        parent_model=target_model,
                    ) or f"pim.attribute.{child_id}.value"
                    product_attribute_ids[target_path] = child_id
                    product_parent_by_attribute_id[child_id] = current_parent_id
                continue
            parent_id = current_parent_id
            for child in child_attributes:
                child_id = int_value(child.get("Id"))
                if child_id is None:
                    continue
                target_path = semantic_pim_field_key(
                    child,
                    parent_attribute=attribute,
                    parent_model=target_model,
                    parent_is_type_series=True,
                ) or f"pim.attribute.{child_id}.value"
                attribute_ids[target_path] = child_id
                parent_by_attribute_id[child_id] = current_parent_id

    if not product_attribute_ids and (parent_id is None or not attribute_ids):
        return DEFAULT_EXPORT_SCHEMA
    return PimExportSchema(
        product_model_id=product_model_id,
        strict_model=True,
        product_attribute_ids=product_attribute_ids,
        product_parent_by_attribute_id=product_parent_by_attribute_id,
        attribute_value_kinds=attribute_value_kinds,
        known_attribute_ids=known_attribute_ids,
        type_series_parent_id=parent_id,
        type_series_attribute_ids=attribute_ids,
        type_series_parent_by_attribute_id=parent_by_attribute_id,
    )


def category_parts(value: Any) -> list[Any]:
    if isinstance(value, (list, tuple, set)):
        result: list[Any] = []
        for item in value:
            result.extend(category_parts(item))
        return result
    if isinstance(value, dict):
        if "value" in value:
            return category_parts(value.get("value"))
        return []
    text = str(value).strip()
    if text.startswith("[") and text.endswith("]"):
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            parsed = None
        if parsed is not None:
            return category_parts(parsed)
    parts = re.split(r"[,;|]", text)
    result = []
    for part in parts:
        clean = part.strip()
        if clean:
            result.append(clean)
    return result


def resolve_product_id(product: dict[str, Any], product_index: dict[str, int]) -> int | None:
    for key in ("product_code", "product_name"):
        value = product.get(key)
        if value in (None, ""):
            continue
        product_id = product_index.get(normalize_lookup(value))
        if product_id is not None:
            return product_id
    return None


def normalize_lookup(value: Any) -> str:
    return normalize_header(value).replace(" ", "")


def parse_number(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def parse_int(value: Any) -> int | None:
    number = parse_number(value)
    return int(number) if number is not None else None


def parse_layer_position(value: Any, fallback: int) -> int:
    number = parse_int(value)
    if number is None:
        return fallback
    return min(max(number, 1), 11)


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = normalize_header(value)
    return text in {"1", "true", "tak", "yes", "y", "domyslny", "domyślny", "x"}


def generated_external_id(name: Any, product_id: int) -> str:
    return stable_hash(str(product_id), str(name)).upper()


def stable_hash(*parts: Any) -> str:
    payload = "|".join(str(part) for part in parts)
    return hashlib.md5(payload.encode("utf-8")).hexdigest()


def make_job_id(filename: str, content: bytes) -> str:
    return stable_hash(filename, content[:2048], len(content))[:12]
