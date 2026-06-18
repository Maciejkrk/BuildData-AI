from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class SourceTable:
    name: str
    rows: list[dict[str, Any]]


def read_source_tables(filename: str, content: bytes) -> list[SourceTable]:
    suffix = Path(filename or "").suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_tables(content)
    if suffix == ".json":
        return read_json_tables(content)
    if suffix in {".csv", ".tsv"}:
        return [SourceTable(Path(filename).stem or "table", read_delimited_rows(content, suffix))]
    raise ValueError("Unsupported file type. Use .xlsx, .json, .csv or .tsv.")


def read_json_tables(content: bytes) -> list[SourceTable]:
    payload = json.loads(content.decode("utf-8-sig"))
    if isinstance(payload, list):
        return [SourceTable("json", [row for row in payload if isinstance(row, dict)])]
    if not isinstance(payload, dict):
        return [SourceTable("json", [])]
    tables: list[SourceTable] = []
    for key, value in payload.items():
        if isinstance(value, list) and all(isinstance(row, dict) for row in value):
            tables.append(SourceTable(str(key), value))
    if not tables:
        tables.append(SourceTable("json", [payload]))
    return tables


def read_xlsx_tables(content: bytes) -> list[SourceTable]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    tables = []
    for sheet in workbook.worksheets:
        rows = rows_from_matrix(list(sheet.iter_rows(values_only=True)))
        if rows:
            tables.append(SourceTable(sheet.title, rows))
    return tables


def read_delimited_rows(content: bytes, suffix: str) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig")
    delimiter = "\t" if suffix == ".tsv" else ","
    return [dict(row) for row in csv.DictReader(StringIO(text), delimiter=delimiter)]


def rows_from_matrix(values: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    if not values:
        return []
    headers = [str(value or "").strip() for value in values[0]]
    rows: list[dict[str, Any]] = []
    for raw_row in values[1:]:
        row = {
            headers[index]: value
            for index, value in enumerate(raw_row)
            if index < len(headers) and headers[index]
        }
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows

